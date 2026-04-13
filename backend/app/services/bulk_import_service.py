"""Service layer for bulk import staging, review, and finalize."""

from __future__ import annotations

import asyncio
import copy
import csv
import hashlib
import io
import json
import multiprocessing
import re
import tempfile
import time
from collections.abc import Callable, Sequence
from datetime import UTC, datetime, timedelta
from multiprocessing.connection import Connection
from pathlib import Path
from typing import Any
from uuid import UUID

import structlog
from fastapi import HTTPException, status
from PyPDF2 import PdfReader
from sqlalchemy import case, delete, func, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.models.bulk_import import ImportItem, ImportRun
from app.models.bulk_import_output import NormalizedLocationDataV1, NormalizedProjectDataV1
from app.models.company import Company
from app.models.discovery_session import DiscoverySession, DiscoverySource
from app.models.intake_suggestion import IntakeSuggestion
from app.models.location import Location
from app.models.project import Project
from app.models.user import User
from app.models.voice_interview import ImportRunIdempotencyKey, VoiceInterview
from app.schemas.bulk_import import (
    BulkImportCompanyResolution,
    BulkImportCompanyResolutionCreateNew,
    BulkImportCompanyResolutionExisting,
    BulkImportFinalizeSummary,
    BulkImportLocationResolution,
    BulkImportLocationResolutionCreateNew,
    BulkImportLocationResolutionExisting,
    BulkImportLocationResolutionLocked,
)
from app.services.bulk_import_ai_extractor import (
    BulkImportAIExtractorError,
    ExtractionDiagnostics,
    ParsedRow,
    bulk_import_ai_extractor,
)
from app.services.idempotency import canonical_sha256
from app.services.intake_field_catalog import build_questionnaire_registry, normalize_field_id
from app.services.s3_service import download_file_content, upload_file_to_s3
from app.services.storage_delete_service import delete_storage_keys
from app.services.voice_status_sync import sync_import_run_status_for_voice
from app.services.voice_transcription_service import voice_transcription_service
from app.services.workspace_service import WORKSPACE_PROJECT_DATA_KEY, WorkspaceService
from app.templates.assessment_questionnaire import get_assessment_questionnaire

logger = structlog.get_logger(__name__)

MAX_PROCESSING_ATTEMPTS = 3
LEASE_SECONDS = 300
RETRY_BASE_SECONDS = 30
RETRY_MAX_SECONDS = 600
RETRY_JITTER_PCT = 0.2
PROCESSING_ERROR_MAX_LENGTH = 500

MAX_IMPORT_FILE_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 2000
MAX_IMPORT_CELLS = 30000
MAX_IMPORT_ITEMS = 4000
PARSER_TIMEOUT_SECONDS = 25
MAX_TEXT_LEN = 4000
PARSER_WORKER_NAME = "bulk-import-parse-worker"

VOICE_CHUNK_TOKENS = 4000
VOICE_CHUNK_OVERLAP_TOKENS = 200
VOICE_MAX_CHUNKS = 20

RETENTION_DAYS_UNFINALIZED = 90
MISSING_LOCATION_REVIEW_NOTE = "Project row missing location context"

ALLOWED_BULK_IMPORT_EXTENSIONS = {
    (ext if ext.startswith(".") else f".{ext}").casefold()
    for ext in settings.bulk_import_allowed_extensions_list
}


class ParserLimitError(ValueError):
    """Raised when parser limits are exceeded."""


def _normalize_token(value: str | None) -> str:
    if not value:
        return ""
    normalized = value.strip().casefold()
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return " ".join(normalized.split())


def _sanitize_text(value: str | None, max_len: int = MAX_TEXT_LEN) -> str | None:
    if value is None:
        return None
    cleaned = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]", "", value)
    cleaned = cleaned.strip()
    if len(cleaned) > max_len:
        return cleaned[:max_len]
    return cleaned


def _sanitize_payload(value: Any) -> Any:
    if isinstance(value, str):
        return _sanitize_text(value)
    if isinstance(value, dict):
        sanitized: dict[str, Any] = {}
        for key, nested_value in value.items():
            key_text = _sanitize_text(str(key), max_len=100)
            if not key_text:
                continue
            sanitized[key_text] = _sanitize_payload(nested_value)
        return sanitized
    if isinstance(value, list):
        return [_sanitize_payload(item) for item in value]
    return value


def _dedupe_backoff_seconds(run_id: UUID, attempt: int) -> int:
    base = min(RETRY_BASE_SECONDS * (2 ** max(attempt - 1, 0)), RETRY_MAX_SECONDS)
    digest = hashlib.sha256(f"{run_id}:{attempt}".encode()).digest()
    jitter_value = int.from_bytes(digest[:8], "big") / 2**64
    jitter_factor = (jitter_value * 2.0 - 1.0) * RETRY_JITTER_PCT
    backoff = round(base * (1 + jitter_factor))
    return max(1, min(backoff, RETRY_MAX_SECONDS))


def _truncate_error(error: str) -> str:
    return error[:PROCESSING_ERROR_MAX_LENGTH]


def _parse_source_subprocess_entrypoint(
    parse_callable: Callable[[str, bytes], list[ParsedRow]],
    filename: str,
    file_bytes: bytes,
    child_conn: Connection,
) -> None:
    """Parse in subprocess, return payload via pipe.

    Payload format:
    - ("ok_file", temp_path)
    - ("error", exc_type_name, message)
    """
    try:
        parsed = parse_callable(filename, file_bytes)
        serialized = _serialize_parsed_rows(parsed)
        with tempfile.NamedTemporaryFile(
            mode="w",
            prefix="bulk-import-parse-",
            suffix=".json",
            encoding="utf-8",
            delete=False,
        ) as temp_file:
            json.dump(serialized, temp_file, separators=(",", ":"))
            temp_path = temp_file.name
        child_conn.send(("ok_file", temp_path))
    except Exception as exc:  # pragma: no cover - validated by parent behavior tests
        child_conn.send(("error", type(exc).__name__, str(exc)))
    finally:
        child_conn.close()


def _default_parse_callable(filename: str, file_bytes: bytes) -> list[ParsedRow]:
    parser_service = BulkImportService()
    return parser_service._parse_source(filename, file_bytes)


def _serialize_parsed_rows(rows: list[ParsedRow]) -> list[dict[str, Any]]:
    serialized: list[dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, ParsedRow):
            raise ValueError("parser_invalid_result")
        serialized.append(
            {
                "location_data": row.location_data,
                "project_data": row.project_data,
                "raw": row.raw,
            }
        )
    return serialized


def _coerce_str_dict(value: Any, *, allow_none: bool) -> dict[str, str] | None:
    if value is None:
        if allow_none:
            return None
        raise ValueError("parser_invalid_result")
    if not isinstance(value, dict):
        raise ValueError("parser_invalid_result")
    coerced: dict[str, str] = {}
    for key, nested in value.items():
        if not isinstance(key, str):
            raise ValueError("parser_invalid_result")
        if nested is None:
            coerced[key] = ""
            continue
        if not isinstance(nested, str):
            raise ValueError("parser_invalid_result")
        coerced[key] = nested
    return coerced


def _deserialize_parsed_rows(payload: Any) -> list[ParsedRow]:
    if not isinstance(payload, list):
        raise ValueError("parser_invalid_result")

    rows: list[ParsedRow] = []
    for item in payload:
        if not isinstance(item, dict):
            raise ValueError("parser_invalid_result")
        location_data = _coerce_str_dict(item.get("location_data"), allow_none=True)
        project_data = _coerce_str_dict(item.get("project_data"), allow_none=True)
        raw_data = _coerce_str_dict(item.get("raw"), allow_none=False)
        if raw_data is None:
            raise ValueError("parser_invalid_result")
        rows.append(
            ParsedRow(
                location_data=location_data,
                project_data=project_data,
                raw=raw_data,
            )
        )
    return rows


def _load_json_parse_result(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as temp_file:
        return json.load(temp_file)


def _log_diagnostics(diagnostics: ExtractionDiagnostics | None) -> dict[str, object]:
    if diagnostics is None:
        return {}
    return {
        "route": diagnostics.route,
        "char_count": diagnostics.char_count,
        "truncated": diagnostics.truncated,
    }


class BulkImportService:
    """Bulk import orchestration across worker and API layers."""

    @staticmethod
    def _ensure_company_active_after_first_stream(company: Company | None) -> None:
        if company is not None and company.account_status == "lead":
            company.account_status = "active"

    def _run_log_context(self, run: ImportRun) -> dict[str, str]:
        return {
            "run_id": str(run.id),
            "organization_id": str(run.organization_id),
            "filename": run.source_filename,
            "source_type": run.source_type,
        }

    def _log_stage_completed(self, run: ImportRun, *, stage: str, started_at: float) -> None:
        logger.info(
            "bulk_import_stage_completed",
            **self._run_log_context(run),
            stage=stage,
            duration_ms=round((time.perf_counter() - started_at) * 1000, 2),
        )

    async def sync_discovery_session_for_run(
        self,
        db: AsyncSession,
        *,
        run_id: UUID,
    ) -> None:
        from app.services.discovery_session_service import DiscoverySessionService

        discovery_service = DiscoverySessionService()
        await discovery_service.sync_session_for_import_run(db, import_run_id=run_id)

    async def _resolve_run_owner_user_id(
        self,
        db: AsyncSession,
        *,
        run: ImportRun,
        fallback_user_id: UUID,
    ) -> UUID:
        discovery_owner_result = await db.execute(
            select(DiscoverySession.assigned_owner_user_id)
            .where(DiscoverySource.import_run_id == run.id)
            .where(DiscoverySource.session_id == DiscoverySession.id)
            .limit(1)
        )
        discovery_owner_user_id = discovery_owner_result.scalar_one_or_none()
        if discovery_owner_user_id is not None:
            return discovery_owner_user_id
        return fallback_user_id

    async def claim_next_run(self, db: AsyncSession) -> ImportRun | None:
        result = await db.execute(
            select(ImportRun)
            .where(ImportRun.status == "uploaded")
            .where(
                (ImportRun.progress_step.is_(None))
                | (ImportRun.progress_step != "discovery_text_pending")
            )
            .where(ImportRun.processing_attempts < MAX_PROCESSING_ATTEMPTS)
            .where(
                (ImportRun.processing_available_at.is_(None))
                | (ImportRun.processing_available_at <= func.now())
            )
            .order_by(ImportRun.processing_available_at.nullsfirst(), ImportRun.created_at)
            .with_for_update(skip_locked=True)
            .limit(1)
        )
        run = result.scalar_one_or_none()
        if run is None:
            return None

        run.processing_attempts += 1
        run.processing_started_at = datetime.now(UTC)
        run.processing_available_at = datetime.now(UTC) + timedelta(seconds=LEASE_SECONDS)
        run.processing_error = None
        run.progress_step = "reading_file"
        if run.source_type == "voice_interview":
            voice = await self._lock_voice_for_run(db, run_id=run.id)
            voice.status = "transcribing"
            voice.error_code = None
            voice.failed_stage = None
            sync_import_run_status_for_voice(run=run, voice_status=voice.status)
        else:
            run.status = "processing"
        return run

    async def requeue_stale_runs(self, db: AsyncSession, limit: int = 100) -> int:
        result = await db.execute(
            select(ImportRun)
            .where(ImportRun.status == "processing")
            .where(
                (ImportRun.progress_step.is_(None))
                | (ImportRun.progress_step != "discovery_text_extracting")
            )
            .where(ImportRun.processing_available_at < func.now())
            .where(ImportRun.processing_attempts < MAX_PROCESSING_ATTEMPTS)
            .order_by(ImportRun.processing_available_at)
            .limit(limit)
            .with_for_update(skip_locked=True)
        )
        runs = list(result.scalars().all())
        for run in runs:
            run.processing_error = "lease_expired_requeued"
            run.processing_available_at = datetime.now(UTC)
            run.processing_started_at = None
            run.progress_step = None
            if run.source_type == "voice_interview":
                voice = await self._lock_voice_for_run(db, run_id=run.id)
                voice.status = "queued"
                sync_import_run_status_for_voice(run=run, voice_status=voice.status)
            else:
                run.status = "uploaded"
        return len(runs)

    async def fail_exhausted_runs(self, db: AsyncSession, limit: int = 100) -> int:
        result = await db.execute(
            select(ImportRun)
            .where(ImportRun.status.in_(["uploaded", "processing"]))
            .where(
                (ImportRun.progress_step.is_(None))
                | (
                    ~ImportRun.progress_step.in_(
                        ["discovery_text_pending", "discovery_text_extracting"]
                    )
                )
            )
            .where(ImportRun.processing_attempts >= MAX_PROCESSING_ATTEMPTS)
            .order_by(ImportRun.updated_at)
            .limit(limit)
            .with_for_update(skip_locked=True)
        )
        runs = list(result.scalars().all())
        for run in runs:
            run.progress_step = None
            run.processing_error = "max_attempts_reached"
            if run.source_type == "voice_interview":
                voice = await self._lock_voice_for_run(db, run_id=run.id)
                voice.status = "failed"
                voice.error_code = "VOICE_MAX_ATTEMPTS_REACHED"
                sync_import_run_status_for_voice(run=run, voice_status=voice.status)
            else:
                run.status = "failed"
            await self.sync_discovery_session_for_run(db, run_id=run.id)
        return len(runs)

    async def purge_expired_artifacts(self, db: AsyncSession, limit: int = 100) -> int:
        cutoff = datetime.now(UTC) - timedelta(days=RETENTION_DAYS_UNFINALIZED)
        result = await db.execute(
            select(ImportRun)
            .where(ImportRun.created_at < cutoff)
            .where(
                ImportRun.status.in_(
                    ["uploaded", "processing", "review_ready", "failed", "no_data"]
                )
            )
            .where(ImportRun.source_type != "voice_interview")
            .where(ImportRun.artifacts_purged_at.is_(None))
            .order_by(ImportRun.created_at)
            .limit(limit)
            .with_for_update(skip_locked=True)
        )
        runs = result.scalars().all()
        if not runs:
            return 0

        purged = 0
        for run in runs:
            try:
                await delete_storage_keys([run.source_file_path])
            except Exception:
                logger.warning(
                    "bulk_import_purge_storage_failed", run_id=str(run.id), exc_info=True
                )
                continue
            run.artifacts_purged_at = datetime.now(UTC)
            run.source_file_path = "imports/purged"
            run.processing_error = None
            await db.execute(
                update(ImportItem)
                .where(ImportItem.run_id == run.id)
                .values(
                    extracted_data={},
                    user_amendments=None,
                    review_notes=None,
                )
            )
            purged += 1
        return purged

    async def process_run(self, db: AsyncSession, run: ImportRun) -> None:
        ai_call_duration_ms: float | None = None
        run_id = run.id
        try:
            if run.status != "processing":
                raise ValueError("run_not_processing")

            if run.source_type == "voice_interview":
                await self._process_voice_run(db=db, run=run)
                return

            await self._persist_progress_checkpoint(db, run, "reading_file")
            read_started = time.perf_counter()
            file_bytes = await download_file_content(run.source_file_path)
            if not file_bytes:
                raise ValueError("empty_file")
            if len(file_bytes) > MAX_IMPORT_FILE_BYTES:
                raise ParserLimitError("max_file_size_exceeded")
            self._log_stage_completed(run, stage="reading_file", started_at=read_started)

            await self._persist_progress_checkpoint(db, run, "identifying_locations")
            identify_started = time.perf_counter()
            extension = Path(run.source_filename).suffix.casefold()
            if extension not in ALLOWED_BULK_IMPORT_EXTENSIONS:
                raise ValueError("unsupported_file_type")
            self._log_stage_completed(
                run,
                stage="identifying_locations",
                started_at=identify_started,
            )

            await self._persist_progress_checkpoint(db, run, "extracting_streams")
            ai_started = time.perf_counter()
            try:
                extraction_result = await bulk_import_ai_extractor.extract_parsed_rows(
                    file_bytes=file_bytes,
                    filename=run.source_filename,
                )
                ai_call_duration_ms = round((time.perf_counter() - ai_started) * 1000, 2)
                logger.info(
                    "bulk_import_bedrock_call_completed",
                    **self._run_log_context(run),
                    status="success",
                    bedrock_duration_ms=ai_call_duration_ms,
                    **_log_diagnostics(extraction_result.diagnostics),
                )
                parsed_rows = extraction_result.rows
            except BulkImportAIExtractorError as exc:
                ai_call_duration_ms = round((time.perf_counter() - ai_started) * 1000, 2)
                logger.info(
                    "bulk_import_bedrock_call_completed",
                    **self._run_log_context(run),
                    status="failed",
                    bedrock_duration_ms=ai_call_duration_ms,
                    error_code=exc.code,
                    **_log_diagnostics(exc.diagnostics),
                )
                raise ValueError(exc.code) from exc

            await self._persist_progress_checkpoint(db, run, "categorizing")
            categorize_started = time.perf_counter()
            await db.execute(delete(ImportItem).where(ImportItem.run_id == run.id))
            staged_items = await self._build_import_items(db, run, parsed_rows)

            if not staged_items:
                run.status = "no_data"
                run.progress_step = None
                run.total_items = 0
                run.accepted_count = 0
                run.rejected_count = 0
                run.amended_count = 0
                run.invalid_count = 0
                run.duplicate_count = 0
                run.processing_error = None
                await self.sync_discovery_session_for_run(db, run_id=run.id)
                await db.flush()
                self._log_stage_completed(run, stage="categorizing", started_at=categorize_started)
                return

            if len(staged_items) > MAX_IMPORT_ITEMS:
                raise ParserLimitError("max_items_exceeded")

            db.add_all(staged_items)
            await db.flush()
            await self.refresh_run_counters(db, run)
            run.status = "review_ready"
            run.progress_step = None
            run.processing_error = None
            await self.sync_discovery_session_for_run(db, run_id=run.id)
            await db.flush()
            self._log_stage_completed(run, stage="categorizing", started_at=categorize_started)
        except Exception as exc:
            await db.rollback()
            await self._handle_processing_failure(db, run_id=run_id, exc=exc)

    async def _process_voice_run(self, db: AsyncSession, run: ImportRun) -> None:
        voice_result = await db.execute(
            select(VoiceInterview)
            .where(VoiceInterview.bulk_import_run_id == run.id)
            .with_for_update()
        )
        voice = voice_result.scalar_one_or_none()
        if voice is None:
            raise ValueError("voice_interview_not_found")

        extension = Path(run.source_filename).suffix.casefold()
        content_type = {
            ".mp3": "audio/mpeg",
            ".wav": "audio/wav",
            ".m4a": "audio/mp4",
        }.get(extension)
        if content_type is None:
            raise ValueError("unsupported_file_type")

        voice.status = "transcribing"
        sync_import_run_status_for_voice(run=run, voice_status=voice.status)
        run.progress_step = "transcribing"
        await db.flush()

        audio_bytes = await download_file_content(run.source_file_path)
        if not audio_bytes:
            raise ValueError("empty_file")

        transcription = await voice_transcription_service.transcribe_audio(
            audio_bytes=audio_bytes,
            filename=run.source_filename,
            content_type=content_type,
        )
        transcript_text = transcription.text.strip()
        if not transcript_text:
            raise ValueError("voice_transcript_empty")

        transcript_key = f"voice-interviews/{run.organization_id}/{voice.id}/transcript.txt"
        await upload_file_to_s3(
            io.BytesIO(transcript_text.encode("utf-8")),
            transcript_key,
            "text/plain",
        )
        voice.transcript_object_key = transcript_key
        run.transcription_model = transcription.model

        voice.status = "extracting"
        sync_import_run_status_for_voice(run=run, voice_status=voice.status)
        run.progress_step = "extracting"
        await db.flush()

        parsed_rows = await self._extract_voice_rows(
            transcript_text=transcript_text,
            source_filename=run.source_filename,
        )
        await db.execute(delete(ImportItem).where(ImportItem.run_id == run.id))
        staged_items = await self._build_import_items(db, run, parsed_rows)

        if not staged_items:
            run.progress_step = None
            run.total_items = 0
            run.accepted_count = 0
            run.rejected_count = 0
            run.amended_count = 0
            run.invalid_count = 0
            run.duplicate_count = 0
            run.processing_error = None
            voice.status = "review_ready"
            sync_import_run_status_for_voice(run=run, voice_status=voice.status)
            await self.sync_discovery_session_for_run(db, run_id=run.id)
            await db.flush()
            return

        if len(staged_items) > MAX_IMPORT_ITEMS:
            raise ParserLimitError("max_items_exceeded")

        db.add_all(staged_items)
        await db.flush()
        await self.refresh_run_counters(db, run)
        voice.status = "review_ready"
        sync_import_run_status_for_voice(run=run, voice_status=voice.status)
        run.progress_step = None
        run.processing_error = None
        voice.error_code = None
        voice.failed_stage = None
        await self.sync_discovery_session_for_run(db, run_id=run.id)
        await db.flush()

    async def _extract_voice_rows(
        self,
        *,
        transcript_text: str,
        source_filename: str,
    ) -> list[ParsedRow]:
        chunks = self._chunk_transcript(transcript_text)
        all_rows: list[ParsedRow] = []
        for index, chunk in enumerate(chunks):
            result = await bulk_import_ai_extractor.extract_parsed_rows_from_text(
                extracted_text=chunk,
                filename=f"voice-chunk-{index + 1}-{source_filename}.txt",
                source_type="voice_interview",
            )
            all_rows.extend(result.rows)

        deduped: dict[str, ParsedRow] = {}
        for row in all_rows:
            location_key = ""
            if row.location_data is not None:
                location_key = self._location_key(
                    self._normalize_location_payload(row.location_data)
                )
            project_name = ""
            if row.project_data is not None:
                project_name = _normalize_token(row.project_data.get("name") or "")
            dedupe_key = f"{location_key}|{project_name}"
            if dedupe_key in deduped:
                continue
            deduped[dedupe_key] = row

        return list(deduped.values())

    def _chunk_transcript(self, transcript_text: str) -> list[str]:
        words = transcript_text.split()
        if not words:
            return []

        max_words = max(1, int(VOICE_CHUNK_TOKENS * 0.75))
        overlap_words = max(0, int(VOICE_CHUNK_OVERLAP_TOKENS * 0.75))
        step = max(1, max_words - overlap_words)

        chunks: list[str] = []
        for start in range(0, len(words), step):
            chunk_words = words[start : start + max_words]
            if not chunk_words:
                continue
            chunks.append(" ".join(chunk_words))
            if start + max_words >= len(words):
                break

        if len(chunks) > VOICE_MAX_CHUNKS:
            raise ValueError("VOICE_TRANSCRIPT_TOO_LONG")
        return chunks

    async def _persist_progress_checkpoint(
        self,
        db: AsyncSession,
        run: ImportRun,
        phase: str,
    ) -> None:
        run.progress_step = phase
        await db.flush()
        await db.commit()

    async def finalize_run(
        self,
        db: AsyncSession,
        *,
        run_id: UUID,
        organization_id: UUID,
        current_user: User,
        resolved_group_ids: list[str] | None = None,
        idempotency_key: str | None = None,
        close_reason: str | None = None,
    ) -> BulkImportFinalizeSummary:
        result = await db.execute(
            select(ImportRun)
            .where(ImportRun.id == run_id, ImportRun.organization_id == organization_id)
            .with_for_update()
        )
        run = result.scalar_one_or_none()
        if not run:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Run not found")

        if run.source_type == "voice_interview":
            return await self._finalize_voice_run(
                db,
                run=run,
                current_user=current_user,
                resolved_group_ids=resolved_group_ids or [],
                idempotency_key=idempotency_key,
                close_reason=close_reason,
            )

        if resolved_group_ids:
            return await self._finalize_bulk_run_subset(
                db,
                run=run,
                current_user=current_user,
                resolved_group_ids=resolved_group_ids,
                close_reason=close_reason,
            )

        if run.status == "completed":
            return self._summary_from_run(run)
        if run.status == "finalizing":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT, detail="Run already finalizing"
            )
        if run.status != "review_ready":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Run is not ready for finalize",
            )
        await self._assert_finalize_ready(db, run)

        run.status = "finalizing"
        run.progress_step = "finalizing"
        run.finalized_by_user_id = current_user.id
        await db.flush()

        all_items_result = await db.execute(
            select(ImportItem).where(ImportItem.run_id == run.id).order_by(ImportItem.created_at)
        )
        all_items = all_items_result.scalars().all()

        active_items = [item for item in all_items if item.status in {"accepted", "amended"}]
        location_items = [item for item in active_items if item.item_type == "location"]
        project_items = [item for item in active_items if item.item_type == "project"]

        await self._assert_finalize_no_new_live_duplicates(
            db=db,
            run=run,
            active_items=active_items,
        )

        company = await self._load_entrypoint_company(db, run)
        company_cache: dict[UUID, Company] = {}
        if company is not None:
            company_cache[company.id] = company
        discovery_provenance = await self._build_discovery_provenance_for_run(db=db, run=run)

        location_by_parent_item_id: dict[UUID, Location] = {}
        created_locations_count = 0
        created_projects_count = 0

        for item in location_items:
            if company is None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Location items are not allowed for location entrypoint",
                )

            selected_existing = await self._location_from_selected_existing_resolution(
                db=db,
                run=run,
                item=item,
                company_id=company.id,
            )
            if selected_existing is not None:
                item.created_location_id = selected_existing.id
                location_by_parent_item_id[item.id] = selected_existing
                continue

            location, created_new = await self._get_or_create_location_for_finalize(
                db=db,
                run=run,
                company_id=company.id,
                current_user=current_user,
                item=item,
            )
            location_by_parent_item_id[item.id] = location
            if created_new:
                created_locations_count += 1

        entrypoint_location: Location | None = None
        if run.entrypoint_type == "location":
            entrypoint_location = await db.get(Location, run.entrypoint_id)
            if (
                not entrypoint_location
                or entrypoint_location.organization_id != run.organization_id
            ):
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND, detail="Entrypoint location not found"
                )

        for item in project_items:
            normalized = self._validated_project_data(item)
            (
                target_location,
                created_location_from_project_resolution,
            ) = await self._resolve_project_location_for_finalize(
                db=db,
                run=run,
                item=item,
                location_by_parent_item_id=location_by_parent_item_id,
                fallback_location=entrypoint_location,
                current_user=current_user,
            )
            if created_location_from_project_resolution:
                created_locations_count += 1
            company_for_project = company_cache.get(target_location.company_id)
            if company_for_project is None:
                company_for_project = await db.get(Company, target_location.company_id)
                if company_for_project:
                    company_cache[target_location.company_id] = company_for_project
            if not company_for_project:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Target location has no company",
                )
            self._ensure_company_active_after_first_stream(company_for_project)

            project_data: dict[str, object] = {
                "technical_sections": copy.deepcopy(get_assessment_questionnaire()),
                WORKSPACE_PROJECT_DATA_KEY: WorkspaceService.build_workspace_v1_seed(
                    material_type=normalized.category,
                    material_name=normalized.name,
                    composition=normalized.description,
                    volume=normalized.volume or normalized.estimated_volume,
                    frequency=normalized.frequency,
                ),
            }
            if normalized.category and normalized.category.strip():
                project_data["bulk_import_category"] = normalized.category.strip()
            self._apply_workspace_provenance(
                project_data=project_data,
                provenance=discovery_provenance,
            )
            project = Project(
                organization_id=run.organization_id,
                user_id=await self._resolve_run_owner_user_id(
                    db,
                    run=run,
                    fallback_user_id=current_user.id,
                ),
                location_id=target_location.id,
                name=normalized.name,
                client=company_for_project.name,
                sector=normalized.sector or company_for_project.sector,
                subsector=normalized.subsector or company_for_project.subsector,
                location=f"{target_location.name}, {target_location.city}",
                project_type=normalized.project_type,
                description=normalized.description,
                budget=0.0,
                schedule_summary="To be defined",
                tags=[],
                status="In Preparation",
                progress=0,
                project_data=project_data,
            )
            db.add(project)
            await db.flush()
            item.created_project_id = project.id
            created_projects_count += 1

        rejected_count = sum(1 for item in all_items if item.status == "rejected")
        invalid_count = sum(1 for item in all_items if item.status == "invalid")
        duplicates_resolved = sum(
            1
            for item in active_items
            if item.duplicate_candidates
            and len(item.duplicate_candidates) > 0
            and item.confirm_create_new
        )

        summary = BulkImportFinalizeSummary(
            run_id=run.id,
            locations_created=created_locations_count,
            projects_created=created_projects_count,
            rejected=rejected_count,
            invalid=invalid_count,
            duplicates_resolved=duplicates_resolved,
        )
        run.status = "completed"
        run.progress_step = None
        run.finalized_at = datetime.now(UTC)
        run.summary_data = summary.model_dump(mode="json")
        await self.refresh_run_counters(db, run)

        if run.source_type != "voice_interview":
            try:
                await delete_storage_keys([run.source_file_path])
                run.artifacts_purged_at = datetime.now(UTC)
            except Exception:
                logger.warning(
                    "bulk_import_finalize_artifact_delete_failed", run_id=str(run.id), exc_info=True
                )

        await db.flush()
        return summary

    async def _finalize_bulk_run_subset(
        self,
        db: AsyncSession,
        *,
        run: ImportRun,
        current_user: User,
        resolved_group_ids: list[str],
        close_reason: str | None,
    ) -> BulkImportFinalizeSummary:
        if close_reason is not None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="close_reason is not supported for non-voice finalize",
            )

        if run.status == "completed":
            return self._summary_from_run(run)
        if run.status == "finalizing":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Run already finalizing",
            )
        if run.status != "review_ready":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Run is not ready for finalize",
            )
        response_summary: BulkImportFinalizeSummary | None = None

        sorted_group_ids = sorted(set(resolved_group_ids))
        if not sorted_group_ids:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="resolved_group_ids required",
            )

        all_items_result = await db.execute(
            select(ImportItem).where(ImportItem.run_id == run.id).with_for_update()
        )
        all_items = list(all_items_result.scalars().all())
        item_index = {entry.id: entry for entry in all_items}
        run_group_ids = sorted(
            {
                group_id
                for group_id in (
                    self._effective_group_id_with_index(item, by_id=item_index)
                    for item in all_items
                )
                if group_id
            }
        )

        unknown_group_ids = [
            group_id for group_id in sorted_group_ids if group_id not in run_group_ids
        ]
        if unknown_group_ids:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail={"message": "Unknown group ids", "group_ids": unknown_group_ids},
            )

        unresolved_in_request = sorted(
            {
                effective_group_id
                for item in all_items
                for effective_group_id in [
                    self._effective_group_id_with_index(item, by_id=item_index)
                ]
                if effective_group_id is not None
                and effective_group_id in sorted_group_ids
                and item.status == "pending_review"
            }
        )
        if unresolved_in_request:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail={
                    "message": "Requested groups are unresolved",
                    "group_ids": unresolved_in_request,
                },
            )

        target_items = [
            item
            for item in all_items
            if self._effective_group_id_with_index(item, by_id=item_index) in sorted_group_ids
        ]
        active_items = [item for item in target_items if item.status in {"accepted", "amended"}]
        location_items = [item for item in active_items if item.item_type == "location"]
        project_items = [item for item in active_items if item.item_type == "project"]

        await self._assert_finalize_no_new_live_duplicates(
            db=db,
            run=run,
            active_items=active_items,
        )

        company = await self._load_entrypoint_company(db, run)
        company_cache: dict[UUID, Company] = {}
        if company is not None:
            company_cache[company.id] = company
        discovery_provenance = await self._build_discovery_provenance_for_run(db=db, run=run)

        location_by_parent_item_id: dict[UUID, Location] = {}
        created_locations_count = 0
        created_projects_count = 0

        for item in location_items:
            if item.created_location_id is not None:
                existing_location = await db.get(Location, item.created_location_id)
                if existing_location is not None:
                    location_by_parent_item_id[item.id] = existing_location
                continue

            if company is None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Location items are not allowed for location entrypoint",
                )

            selected_existing = await self._location_from_selected_existing_resolution(
                db=db,
                run=run,
                item=item,
                company_id=company.id,
            )
            if selected_existing is not None:
                item.created_location_id = selected_existing.id
                location_by_parent_item_id[item.id] = selected_existing
                continue

            location, created_new = await self._get_or_create_location_for_finalize(
                db=db,
                run=run,
                company_id=company.id,
                current_user=current_user,
                item=item,
            )
            location_by_parent_item_id[item.id] = location
            if created_new:
                created_locations_count += 1

        entrypoint_location: Location | None = None
        if run.entrypoint_type == "location":
            entrypoint_location = await db.get(Location, run.entrypoint_id)
            if (
                not entrypoint_location
                or entrypoint_location.organization_id != run.organization_id
            ):
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Entrypoint location not found",
                )

        for item in project_items:
            if item.created_project_id is not None:
                continue

            normalized = self._validated_project_data(item)
            (
                target_location,
                created_location_from_project_resolution,
            ) = await self._resolve_project_location_for_finalize(
                db=db,
                run=run,
                item=item,
                location_by_parent_item_id=location_by_parent_item_id,
                fallback_location=entrypoint_location,
                current_user=current_user,
            )
            if created_location_from_project_resolution:
                created_locations_count += 1
            company_for_project = company_cache.get(target_location.company_id)
            if company_for_project is None:
                company_for_project = await db.get(Company, target_location.company_id)
                if company_for_project is not None:
                    company_cache[target_location.company_id] = company_for_project
            if company_for_project is None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Target location has no company",
                )
            self._ensure_company_active_after_first_stream(company_for_project)

            project_data: dict[str, object] = {
                "technical_sections": copy.deepcopy(get_assessment_questionnaire()),
                WORKSPACE_PROJECT_DATA_KEY: WorkspaceService.build_workspace_v1_seed(
                    material_type=normalized.category,
                    material_name=normalized.name,
                    composition=normalized.description,
                    volume=normalized.volume or normalized.estimated_volume,
                    frequency=normalized.frequency,
                ),
            }
            if normalized.category and normalized.category.strip():
                project_data["bulk_import_category"] = normalized.category.strip()
            self._apply_workspace_provenance(
                project_data=project_data,
                provenance=discovery_provenance,
            )

            project = Project(
                organization_id=run.organization_id,
                user_id=await self._resolve_run_owner_user_id(
                    db,
                    run=run,
                    fallback_user_id=current_user.id,
                ),
                location_id=target_location.id,
                name=normalized.name,
                client=company_for_project.name,
                sector=normalized.sector or company_for_project.sector,
                subsector=normalized.subsector or company_for_project.subsector,
                location=f"{target_location.name}, {target_location.city}",
                project_type=normalized.project_type,
                description=normalized.description,
                budget=0.0,
                schedule_summary="To be defined",
                tags=[],
                status="In Preparation",
                progress=0,
                project_data=project_data,
            )
            db.add(project)
            await db.flush()
            item.created_project_id = project.id
            created_projects_count += 1

        previously_counted_groups = self._counted_summary_groups(run)
        new_group_ids = set(sorted_group_ids) - previously_counted_groups
        newly_counted_target_items = [
            item
            for item in target_items
            if self._effective_group_id_with_index(item, by_id=item_index) in new_group_ids
        ]
        newly_counted_active_items = [
            item for item in active_items if item in newly_counted_target_items
        ]

        rejected_count = sum(1 for item in newly_counted_target_items if item.status == "rejected")
        invalid_count = sum(1 for item in newly_counted_target_items if item.status == "invalid")
        duplicates_resolved = sum(
            1
            for item in newly_counted_active_items
            if item.duplicate_candidates
            and len(item.duplicate_candidates) > 0
            and item.confirm_create_new
        )

        delta_summary = BulkImportFinalizeSummary(
            run_id=run.id,
            locations_created=created_locations_count,
            projects_created=created_projects_count,
            rejected=rejected_count,
            invalid=invalid_count,
            duplicates_resolved=duplicates_resolved,
        )
        response_summary = self._accumulate_run_summary(
            run=run,
            delta=delta_summary,
        )

        unresolved_group_count = len(
            {
                effective_group_id
                for item in all_items
                for effective_group_id in [
                    self._effective_group_id_with_index(item, by_id=item_index)
                ]
                if effective_group_id is not None
                and (
                    item.status == "pending_review"
                    or (
                        item.status in {"accepted", "amended"}
                        and item.item_type == "location"
                        and item.created_location_id is None
                    )
                    or (
                        item.status in {"accepted", "amended"}
                        and item.item_type == "project"
                        and item.created_project_id is None
                    )
                )
            }
        )

        run.finalized_by_user_id = current_user.id
        run.progress_step = None
        if unresolved_group_count == 0:
            run.status = "completed"
            run.finalized_at = datetime.now(UTC)
            try:
                await delete_storage_keys([run.source_file_path])
                run.artifacts_purged_at = datetime.now(UTC)
            except Exception:
                logger.warning(
                    "bulk_import_finalize_artifact_delete_failed",
                    run_id=str(run.id),
                    exc_info=True,
                )
        else:
            run.status = "review_ready"

        summary_payload = response_summary.model_dump(mode="json")
        summary_payload["countedGroupIds"] = sorted(
            self._counted_summary_groups(run).union(sorted_group_ids)
        )
        run.summary_data = summary_payload

        await self.refresh_run_counters(db, run)
        await db.flush()
        if response_summary is None:
            raise RuntimeError("subset_finalize_summary_missing")
        return response_summary

    def _accumulate_run_summary(
        self,
        *,
        run: ImportRun,
        delta: BulkImportFinalizeSummary,
    ) -> BulkImportFinalizeSummary:
        previous = self._summary_from_run(run) if run.summary_data else None
        if previous is None:
            return BulkImportFinalizeSummary(
                run_id=run.id,
                locations_created=delta.locations_created,
                projects_created=delta.projects_created,
                rejected=delta.rejected,
                invalid=delta.invalid,
                duplicates_resolved=delta.duplicates_resolved,
            )
        return BulkImportFinalizeSummary(
            run_id=run.id,
            locations_created=previous.locations_created + delta.locations_created,
            projects_created=previous.projects_created + delta.projects_created,
            rejected=previous.rejected + delta.rejected,
            invalid=previous.invalid + delta.invalid,
            duplicates_resolved=previous.duplicates_resolved + delta.duplicates_resolved,
        )

    def _counted_summary_groups(self, run: ImportRun) -> set[str]:
        if not isinstance(run.summary_data, dict):
            return set()
        raw = run.summary_data.get("countedGroupIds")
        if not isinstance(raw, list):
            return set()
        return {group_id for group_id in raw if isinstance(group_id, str) and group_id}

    async def _assert_finalize_no_new_live_duplicates(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        active_items: Sequence[ImportItem],
    ) -> None:
        if not active_items:
            return

        company: Company | None = None
        if run.entrypoint_type == "company":
            company = await self._load_entrypoint_company(db, run)

        entrypoint_location: Location | None = None
        if run.entrypoint_type == "location":
            entrypoint_location = await db.get(Location, run.entrypoint_id)
            if (
                not entrypoint_location
                or entrypoint_location.organization_id != run.organization_id
            ):
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Entrypoint location not found",
                )

        active_items_by_id = {item.id: item for item in active_items}

        for item in active_items:
            if item.confirm_create_new:
                continue

            if item.item_type == "location":
                if self._selected_existing_location_id(item) is not None:
                    continue
                if company is None:
                    continue
                normalized_location = self._validated_location_data(item)
                location_duplicate_exists = await self._location_duplicate_exists_for_finalize(
                    db=db,
                    organization_id=run.organization_id,
                    company_id=company.id,
                    location_data=normalized_location,
                )
                if location_duplicate_exists:
                    raise HTTPException(
                        status_code=status.HTTP_409_CONFLICT,
                        detail=f"Duplicate detected before finalize for item {item.id}",
                    )
                continue

            normalized_project = self._validated_project_data(item)
            target_location_id = await self._project_finalize_location_id_for_duplicate_recheck(
                db=db,
                run=run,
                item=item,
                active_items_by_id=active_items_by_id,
                entrypoint_location=entrypoint_location,
                company=company,
            )
            if target_location_id is None:
                continue

            project_duplicate_exists = await self._project_duplicate_exists_for_finalize(
                db=db,
                organization_id=run.organization_id,
                location_id=target_location_id,
                project_name=normalized_project.name,
            )
            if project_duplicate_exists:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"Duplicate detected before finalize for item {item.id}",
                )

    async def _finalize_voice_run(
        self,
        db: AsyncSession,
        *,
        run: ImportRun,
        current_user: User,
        resolved_group_ids: list[str],
        idempotency_key: str | None,
        close_reason: str | None,
    ) -> BulkImportFinalizeSummary:
        voice_result = await db.execute(
            select(VoiceInterview)
            .where(VoiceInterview.bulk_import_run_id == run.id)
            .with_for_update()
        )
        voice = voice_result.scalar_one_or_none()
        if voice is None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT, detail="Voice interview not found"
            )

        sorted_group_ids = sorted(set(resolved_group_ids))
        request_hash = canonical_sha256(
            {
                "resolved_group_ids_sorted": sorted_group_ids,
                "close_reason": close_reason,
                "run_id": str(run.id),
            }
        )

        if close_reason != "empty_extraction" and not idempotency_key:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="idempotency_key required"
            )

        idempotency_record: ImportRunIdempotencyKey | None = None
        if idempotency_key:
            idem_result = await db.execute(
                select(ImportRunIdempotencyKey)
                .where(
                    ImportRunIdempotencyKey.operation_type == "finalize",
                    ImportRunIdempotencyKey.run_id == run.id,
                    ImportRunIdempotencyKey.idempotency_key == idempotency_key,
                )
                .with_for_update()
            )
            idempotency_record = idem_result.scalar_one_or_none()
            if idempotency_record is not None:
                if idempotency_record.request_hash != request_hash:
                    raise HTTPException(
                        status_code=status.HTTP_409_CONFLICT,
                        detail={
                            "message": "Idempotency key payload mismatch",
                            "code": "IDEMPOTENCY_KEY_PAYLOAD_MISMATCH",
                        },
                    )
                persisted_summary = idempotency_record.response_json.get("summary")
                if not isinstance(persisted_summary, dict):
                    raise HTTPException(
                        status_code=status.HTTP_409_CONFLICT,
                        detail="Invalid idempotency replay payload",
                    )
                return BulkImportFinalizeSummary.model_validate(persisted_summary)

        if run.status != "review_ready" or voice.status not in {
            "review_ready",
            "partial_finalized",
        }:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Voice run is not ready for finalize",
            )

        items_result = await db.execute(
            select(ImportItem).where(ImportItem.run_id == run.id).with_for_update()
        )
        all_items = list(items_result.scalars().all())
        item_index = {entry.id: entry for entry in all_items}
        run_group_ids = sorted(
            {
                group_id
                for group_id in (
                    self._effective_group_id_with_index(item, by_id=item_index)
                    for item in all_items
                )
                if group_id
            }
        )

        if close_reason == "empty_extraction":
            if sorted_group_ids:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="resolved_group_ids must be empty for empty extraction close",
                )
            if run_group_ids:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="empty extraction close only allowed when run has zero groups",
                )
            summary = BulkImportFinalizeSummary(
                run_id=run.id,
                locations_created=0,
                projects_created=0,
                rejected=0,
                invalid=0,
                duplicates_resolved=0,
            )
            voice.status = "finalized"
            sync_import_run_status_for_voice(run=run, voice_status=voice.status)
            run.finalized_by_user_id = current_user.id
            run.finalized_at = datetime.now(UTC)
            run.summary_data = summary.model_dump(mode="json")
            if idempotency_key and idempotency_record is None:
                db.add(
                    ImportRunIdempotencyKey(
                        operation_type="finalize",
                        run_id=run.id,
                        idempotency_key=idempotency_key,
                        request_hash=request_hash,
                        response_json={
                            "status": run.status,
                            "summary": summary.model_dump(mode="json"),
                        },
                        response_status_code=200,
                    )
                )
            await db.flush()
            return summary

        if not sorted_group_ids:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="resolved_group_ids required",
            )

        unknown_group_ids = [
            group_id for group_id in sorted_group_ids if group_id not in run_group_ids
        ]
        if unknown_group_ids:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail={"message": "Unknown group ids", "group_ids": unknown_group_ids},
            )

        unresolved_in_request = sorted(
            {
                effective_group_id
                for item in all_items
                for effective_group_id in [
                    self._effective_group_id_with_index(item, by_id=item_index)
                ]
                if effective_group_id is not None
                and effective_group_id in sorted_group_ids
                and item.status == "pending_review"
            }
        )
        if unresolved_in_request:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail={
                    "message": "Requested groups are unresolved",
                    "group_ids": unresolved_in_request,
                },
            )

        target_items = [
            item
            for item in all_items
            if self._effective_group_id_with_index(item, by_id=item_index) in sorted_group_ids
        ]
        active_items = [item for item in target_items if item.status in {"accepted", "amended"}]
        location_items = [item for item in active_items if item.item_type == "location"]
        project_items = [item for item in active_items if item.item_type == "project"]

        company = await self._load_entrypoint_company(db, run)
        company_cache: dict[UUID, Company] = {}
        if company is not None:
            company_cache[company.id] = company
        discovery_provenance = await self._build_discovery_provenance_for_run(db=db, run=run)

        location_by_parent_item_id: dict[UUID, Location] = {}
        created_locations_count = 0
        created_projects_count = 0

        for item in location_items:
            if item.created_location_id is not None:
                existing_location = await db.get(Location, item.created_location_id)
                if existing_location is not None:
                    location_by_parent_item_id[item.id] = existing_location
                continue

            if company is None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Location items are not allowed for location entrypoint",
                )

            selected_existing = await self._location_from_selected_existing_resolution(
                db=db,
                run=run,
                item=item,
                company_id=company.id,
            )
            if selected_existing is not None:
                item.created_location_id = selected_existing.id
                location_by_parent_item_id[item.id] = selected_existing
                continue

            mapped_location = await self._resolve_existing_location_for_finalize(
                db=db,
                run=run,
                item=item,
                company=company,
            )
            if mapped_location is not None:
                item.created_location_id = mapped_location.id
                location_by_parent_item_id[item.id] = mapped_location
                continue

            location, created_new = await self._get_or_create_location_for_finalize(
                db=db,
                run=run,
                company_id=company.id,
                current_user=current_user,
                item=item,
            )
            location_by_parent_item_id[item.id] = location
            if created_new:
                created_locations_count += 1

        entrypoint_location: Location | None = None
        if run.entrypoint_type == "location":
            entrypoint_location = await db.get(Location, run.entrypoint_id)
            if (
                not entrypoint_location
                or entrypoint_location.organization_id != run.organization_id
            ):
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Entrypoint location not found",
                )

        registry = build_questionnaire_registry()
        suggestions_created_count = 0

        for item in project_items:
            if item.created_project_id is not None:
                continue
            normalized = self._validated_project_data(item)
            (
                target_location,
                created_location_from_project_resolution,
            ) = await self._resolve_project_location_for_finalize(
                db=db,
                run=run,
                item=item,
                location_by_parent_item_id=location_by_parent_item_id,
                fallback_location=entrypoint_location,
                current_user=current_user,
            )
            if created_location_from_project_resolution:
                created_locations_count += 1
            mapped_project = await self._resolve_existing_project_for_finalize(
                db=db,
                run=run,
                item=item,
                target_location=target_location,
            )
            if mapped_project is not None:
                item.created_project_id = mapped_project.id
                continue
            company_for_project = company_cache.get(target_location.company_id)
            if company_for_project is None:
                company_for_project = await db.get(Company, target_location.company_id)
                if company_for_project is not None:
                    company_cache[target_location.company_id] = company_for_project
            if company_for_project is None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Target location has no company",
                )

            project_data: dict[str, object] = {
                "technical_sections": copy.deepcopy(get_assessment_questionnaire()),
                WORKSPACE_PROJECT_DATA_KEY: WorkspaceService.build_workspace_v1_seed(
                    material_type=normalized.category,
                    material_name=normalized.name,
                    composition=normalized.description,
                    volume=normalized.volume or normalized.estimated_volume,
                    frequency=normalized.frequency,
                ),
            }
            if normalized.category and normalized.category.strip():
                project_data["bulk_import_category"] = normalized.category.strip()
            self._apply_workspace_provenance(
                project_data=project_data,
                provenance=discovery_provenance,
            )
            self._ensure_company_active_after_first_stream(company_for_project)

            project = Project(
                organization_id=run.organization_id,
                user_id=await self._resolve_run_owner_user_id(
                    db,
                    run=run,
                    fallback_user_id=current_user.id,
                ),
                location_id=target_location.id,
                name=normalized.name,
                client=company_for_project.name,
                sector=normalized.sector or company_for_project.sector,
                subsector=normalized.subsector or company_for_project.subsector,
                location=f"{target_location.name}, {target_location.city}",
                project_type=normalized.project_type,
                description=normalized.description,
                budget=0.0,
                schedule_summary="To be defined",
                tags=[],
                status="In Preparation",
                progress=0,
                project_data=project_data,
            )
            db.add(project)
            await db.flush()
            item.created_project_id = project.id
            created_projects_count += 1
            suggestions_created_count += self._create_pending_intake_suggestions_from_project_item(
                db=db,
                project=project,
                source_item=item,
                registry=registry,
                user_id=current_user.id,
            )

        if suggestions_created_count > 0:
            logger.info(
                "voice_finalize_intake_suggestions_created",
                run_id=str(run.id),
                suggestions_created=suggestions_created_count,
            )

        unresolved_group_count = len(
            {
                effective_group_id
                for item in all_items
                for effective_group_id in [
                    self._effective_group_id_with_index(item, by_id=item_index)
                ]
                if effective_group_id is not None
                and (
                    item.status == "pending_review"
                    or (
                        item.status in {"accepted", "amended"}
                        and item.item_type == "location"
                        and item.created_location_id is None
                    )
                    or (
                        item.status in {"accepted", "amended"}
                        and item.item_type == "project"
                        and item.created_project_id is None
                    )
                )
            }
        )

        rejected_count = sum(1 for item in target_items if item.status == "rejected")
        invalid_count = sum(1 for item in target_items if item.status == "invalid")
        duplicates_resolved = sum(
            1
            for item in active_items
            if item.duplicate_candidates
            and len(item.duplicate_candidates) > 0
            and item.confirm_create_new
        )

        summary = BulkImportFinalizeSummary(
            run_id=run.id,
            locations_created=created_locations_count,
            projects_created=created_projects_count,
            rejected=rejected_count,
            invalid=invalid_count,
            duplicates_resolved=duplicates_resolved,
        )

        voice.status = "finalized" if unresolved_group_count == 0 else "partial_finalized"
        sync_import_run_status_for_voice(run=run, voice_status=voice.status)
        run.finalized_by_user_id = current_user.id
        if voice.status == "finalized":
            run.finalized_at = datetime.now(UTC)
        run.summary_data = summary.model_dump(mode="json")
        await self.refresh_run_counters(db, run)

        if idempotency_key and idempotency_record is None:
            db.add(
                ImportRunIdempotencyKey(
                    operation_type="finalize",
                    run_id=run.id,
                    idempotency_key=idempotency_key,
                    request_hash=request_hash,
                    response_json={
                        "status": run.status,
                        "summary": summary.model_dump(mode="json"),
                    },
                    response_status_code=200,
                )
            )

        await db.flush()
        return summary

    async def _resolve_existing_location_for_finalize(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        item: ImportItem,
        company: Company | None,
    ) -> Location | None:
        if item.confirm_create_new:
            return None
        candidate_ids = self._duplicate_candidate_ids(item)
        if not candidate_ids:
            return None
        if len(candidate_ids) > 1:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Location duplicate mapping ambiguous for item {item.id}",
            )
        if company is None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Company entrypoint required",
            )
        location = await db.get(Location, candidate_ids[0])
        if (
            location is None
            or location.organization_id != run.organization_id
            or location.company_id != company.id
        ):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Location duplicate target invalid for item {item.id}",
            )
        return location

    async def _resolve_existing_project_for_finalize(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        item: ImportItem,
        target_location: Location,
    ) -> Project | None:
        if item.confirm_create_new:
            return None
        candidate_ids = self._duplicate_candidate_ids(item)
        if not candidate_ids:
            return None
        if len(candidate_ids) > 1:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Project duplicate mapping ambiguous for item {item.id}",
            )
        project = await db.get(Project, candidate_ids[0])
        if (
            project is None
            or project.organization_id != run.organization_id
            or project.location_id != target_location.id
        ):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Project duplicate target invalid for item {item.id}",
            )
        return project

    def _selected_existing_location_id(self, item: ImportItem) -> UUID | None:
        amendments = item.user_amendments
        if not isinstance(amendments, dict):
            return None
        resolution = amendments.get("location_resolution")
        if not isinstance(resolution, dict):
            return None
        mode_raw: object | None = None
        location_id_raw: object | None = None
        for key, value in resolution.items():
            if key == "mode":
                mode_raw = value
            if key == "location_id":
                location_id_raw = value
        if mode_raw != "existing":
            return None
        if not isinstance(location_id_raw, str):
            return None
        try:
            return UUID(location_id_raw)
        except ValueError:
            return None

    def _location_resolution_mode(self, item: ImportItem) -> str | None:
        amendments = item.user_amendments
        if not isinstance(amendments, dict):
            return None
        resolution = amendments.get("location_resolution")
        if not isinstance(resolution, dict):
            return None
        mode_raw: object | None = None
        for key, value in resolution.items():
            if key == "mode":
                mode_raw = value
                break
        if isinstance(mode_raw, str):
            return mode_raw
        return None

    async def _location_from_selected_existing_resolution(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        item: ImportItem,
        company_id: UUID,
    ) -> Location | None:
        selected_location_id = self._selected_existing_location_id(item)
        if selected_location_id is None:
            return None
        selected_location = await db.get(Location, selected_location_id)
        if (
            selected_location is None
            or selected_location.organization_id != run.organization_id
            or selected_location.company_id != company_id
        ):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Selected location invalid for item {item.id}",
            )
        return selected_location

    async def _find_location_by_identity(
        self,
        *,
        db: AsyncSession,
        organization_id: UUID,
        company_id: UUID,
        location_data: NormalizedLocationDataV1,
    ) -> Location | None:
        result = await db.execute(
            select(Location)
            .where(
                Location.organization_id == organization_id,
                Location.company_id == company_id,
                func.lower(Location.name) == location_data.name.casefold(),
                func.lower(Location.city) == location_data.city.casefold(),
                func.lower(Location.state) == location_data.state.casefold(),
            )
            .limit(1)
        )
        return result.scalar_one_or_none()

    async def _get_or_create_location_for_finalize(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        company_id: UUID,
        current_user: User,
        item: ImportItem,
    ) -> tuple[Location, bool]:
        normalized = self._validated_location_data(item)
        resolution_mode = self._location_resolution_mode(item)
        existing_location = await self._find_location_by_identity(
            db=db,
            organization_id=run.organization_id,
            company_id=company_id,
            location_data=normalized,
        )
        if existing_location is not None:
            if resolution_mode == "create_new":
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Location already exists. Use existing location instead.",
                )
            item.created_location_id = existing_location.id
            return existing_location, False

        location = Location(
            organization_id=run.organization_id,
            company_id=company_id,
            name=normalized.name,
            city=normalized.city,
            state=normalized.state,
            address=normalized.address,
            created_by_user_id=current_user.id,
        )
        db.add(location)
        await db.flush()
        item.created_location_id = location.id
        return location, True

    def _duplicate_candidate_ids(self, item: ImportItem) -> list[UUID]:
        candidates = item.duplicate_candidates or []
        ids: list[UUID] = []
        for candidate in candidates:
            raw_id = candidate.get("id") if isinstance(candidate, dict) else None
            if not isinstance(raw_id, str):
                continue
            try:
                parsed_id = UUID(raw_id)
            except ValueError:
                continue
            ids.append(parsed_id)
        return list(dict.fromkeys(ids))

    async def _lock_voice_for_run(self, db: AsyncSession, *, run_id: UUID) -> VoiceInterview:
        result = await db.execute(
            select(VoiceInterview)
            .where(VoiceInterview.bulk_import_run_id == run_id)
            .with_for_update()
        )
        voice = result.scalar_one_or_none()
        if voice is None:
            raise ValueError("voice_interview_not_found")
        return voice

    async def _location_duplicate_exists_for_finalize(
        self,
        *,
        db: AsyncSession,
        organization_id: UUID,
        company_id: UUID,
        location_data: NormalizedLocationDataV1,
    ) -> bool:
        result = await db.execute(
            select(Location.id)
            .where(
                Location.organization_id == organization_id,
                Location.company_id == company_id,
                func.lower(Location.name) == location_data.name.casefold(),
                func.lower(Location.city) == location_data.city.casefold(),
                func.lower(Location.state) == location_data.state.casefold(),
            )
            .limit(1)
        )
        return result.scalar_one_or_none() is not None

    async def _project_duplicate_exists_for_finalize(
        self,
        *,
        db: AsyncSession,
        organization_id: UUID,
        location_id: UUID,
        project_name: str,
    ) -> bool:
        normalized_project_name = _sanitize_text(project_name)
        if not normalized_project_name:
            return False
        result = await db.execute(
            select(Project.id)
            .where(
                Project.organization_id == organization_id,
                Project.location_id == location_id,
                func.lower(Project.name) == normalized_project_name.casefold(),
            )
            .limit(1)
        )
        return result.scalar_one_or_none() is not None

    async def _project_finalize_location_id_for_duplicate_recheck(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        item: ImportItem,
        active_items_by_id: dict[UUID, ImportItem],
        entrypoint_location: Location | None,
        company: Company | None,
    ) -> UUID | None:
        if run.entrypoint_type == "location":
            if entrypoint_location is None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Entrypoint location missing",
                )
            return entrypoint_location.id

        if item.parent_item_id is not None:
            parent_item = active_items_by_id.get(item.parent_item_id)
            if parent_item is not None:
                return None

        selected_existing_location_id = self._selected_existing_location_id(item)
        if selected_existing_location_id is not None:
            return selected_existing_location_id

        if company is not None:
            create_new_location_data = self._orphan_project_create_new_location_data(item)
            if create_new_location_data is not None:
                existing_location = await self._find_location_by_identity(
                    db=db,
                    organization_id=run.organization_id,
                    company_id=company.id,
                    location_data=create_new_location_data,
                )
                if existing_location is not None:
                    return existing_location.id

        location_name = _sanitize_text(str(item.normalized_data.get("location_name") or ""))
        city = _sanitize_text(str(item.normalized_data.get("location_city") or ""))
        state_value = _sanitize_text(str(item.normalized_data.get("location_state") or ""))
        if not location_name or not city or not state_value:
            return None

        if company is None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Company entrypoint required",
            )

        result = await db.execute(
            select(Location.id)
            .where(
                Location.organization_id == run.organization_id,
                Location.company_id == company.id,
                func.lower(Location.name) == location_name.casefold(),
                func.lower(Location.city) == city.casefold(),
                func.lower(Location.state) == state_value.casefold(),
            )
            .limit(1)
        )
        return result.scalar_one_or_none()

    async def refresh_run_counters(self, db: AsyncSession, run: ImportRun) -> None:
        counts_result = await db.execute(
            select(
                func.count(ImportItem.id),
                func.sum(case((ImportItem.status == "accepted", 1), else_=0)),
                func.sum(case((ImportItem.status == "rejected", 1), else_=0)),
                func.sum(case((ImportItem.status == "amended", 1), else_=0)),
                func.sum(case((ImportItem.status == "invalid", 1), else_=0)),
                func.sum(
                    case(
                        (
                            func.jsonb_typeof(ImportItem.duplicate_candidates) == "array",
                            case(
                                (func.jsonb_array_length(ImportItem.duplicate_candidates) > 0, 1),
                                else_=0,
                            ),
                        ),
                        else_=0,
                    )
                ),
            ).where(ImportItem.run_id == run.id)
        )
        counts = counts_result.one()
        run.total_items = int(counts[0] or 0)
        run.accepted_count = int(counts[1] or 0)
        run.rejected_count = int(counts[2] or 0)
        run.amended_count = int(counts[3] or 0)
        run.invalid_count = int(counts[4] or 0)
        run.duplicate_count = int(counts[5] or 0)

    async def decide_discovery_project_draft(
        self,
        db: AsyncSession,
        *,
        organization_id: UUID,
        item_id: UUID,
        current_user: User,
        action: str,
        normalized_data: dict[str, object] | None,
        review_notes: str | None,
        company_resolution: BulkImportCompanyResolution | None,
        location_resolution: BulkImportLocationResolution | None,
        confirm_create_new: bool | None,
        owner_user_id: UUID | None,
    ) -> tuple[ImportItem, BulkImportFinalizeSummary, ImportRun]:
        run_result = await db.execute(
            select(ImportRun)
            .join(ImportItem, ImportItem.run_id == ImportRun.id)
            .where(ImportItem.id == item_id, ImportItem.organization_id == organization_id)
            .where(ImportRun.organization_id == organization_id)
            .with_for_update()
        )
        run = run_result.scalar_one_or_none()
        if run is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
        discovery_source_type_result = await db.execute(
            select(DiscoverySource).where(DiscoverySource.import_run_id == run.id).limit(1)
        )
        discovery_source = discovery_source_type_result.scalar_one_or_none()
        discovery_source_type = discovery_source.source_type if discovery_source else None
        if discovery_source_type not in {"file", "text", "audio"}:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Discovery decision endpoint is only available for discovery runs",
            )

        result = await db.execute(
            select(ImportItem)
            .where(ImportItem.id == item_id, ImportItem.organization_id == organization_id)
            .where(ImportItem.run_id == run.id)
            .with_for_update()
        )
        item = result.scalar_one_or_none()
        if item is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
        if item.item_type != "project":
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Discovery decision endpoint only supports project drafts",
            )
        if run.status != "review_ready":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Run is not editable",
            )

        resolved_owner_user_id = await self._resolve_discovery_owner_user_id(
            db,
            run=run,
            current_user=current_user,
            owner_user_id=owner_user_id,
        )
        if item.created_project_id is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Draft already resolved",
            )
        if item.status in {"rejected", "invalid"}:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Draft already terminal",
            )

        summary = BulkImportFinalizeSummary(
            run_id=run.id,
            locations_created=0,
            projects_created=0,
            rejected=0,
            invalid=0,
            duplicates_resolved=0,
        )

        if action == "reject":
            item.status = "rejected"
            item.needs_review = False
            summary = BulkImportFinalizeSummary(
                run_id=run.id,
                locations_created=0,
                projects_created=0,
                rejected=1,
                invalid=0,
                duplicates_resolved=0,
            )
        elif action == "confirm":
            if normalized_data is None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="normalized_data required for confirm",
                )
            if review_notes is not None:
                sanitized_review_notes = _sanitize_text(review_notes, max_len=1000)
                item.review_notes = sanitized_review_notes or None
            sanitized_patch = _sanitize_payload(normalized_data)
            merged = dict(item.normalized_data)
            merged.update(sanitized_patch)
            item.normalized_data = merged
            existing_user_amendments = (
                item.user_amendments if isinstance(item.user_amendments, dict) else {}
            )
            merged_user_amendments = dict(existing_user_amendments)
            merged_user_amendments.update(sanitized_patch)
            item.user_amendments = merged_user_amendments
            if company_resolution is not None:
                await self._apply_company_resolution(
                    db=db,
                    run=run,
                    item=item,
                    company_resolution=company_resolution,
                )
            resolved_company_for_run = await self._resolve_company_for_discovery_draft(
                db=db,
                run=run,
                item=item,
                current_user=current_user,
            )
            if location_resolution is not None:
                await self._apply_location_resolution(
                    db=db,
                    run=run,
                    item=item,
                    location_resolution=location_resolution,
                    effective_company_id_override=(
                        resolved_company_for_run.id
                        if resolved_company_for_run is not None
                        else None
                    ),
                )
            if confirm_create_new is not None:
                item.confirm_create_new = confirm_create_new
            item.status = "amended"
            item.needs_review = self._needs_review(item.item_type, item.normalized_data)
            self._ensure_duplicate_confirmation(item, target_status="amended")
            normalized = self._validated_project_data(item)
            company = resolved_company_for_run
            if company is None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Company entrypoint required",
                )
            location_by_parent_item_id: dict[UUID, Location] = {}
            treat_as_orphan = item.parent_item_id is None
            if item.parent_item_id is not None:
                parent_item = await db.get(ImportItem, item.parent_item_id)
                if parent_item is None:
                    treat_as_orphan = True
                else:
                    if parent_item.created_location_id is not None:
                        created_parent_location = await db.get(
                            Location, parent_item.created_location_id
                        )
                        if created_parent_location is None:
                            raise HTTPException(
                                status_code=status.HTTP_409_CONFLICT,
                                detail="Parent location mapping missing",
                            )
                        location_by_parent_item_id[parent_item.id] = created_parent_location
                    if location_resolution is not None:
                        await self._apply_location_resolution(
                            db=db,
                            run=run,
                            item=parent_item,
                            location_resolution=location_resolution,
                            effective_company_id_override=company.id,
                        )
                    if parent_item.id not in location_by_parent_item_id:
                        selected_existing = await self._location_from_selected_existing_resolution(
                            db=db,
                            run=run,
                            item=parent_item,
                            company_id=company.id,
                        )
                        if selected_existing is not None:
                            parent_item.created_location_id = selected_existing.id
                            location_by_parent_item_id[parent_item.id] = selected_existing
                        else:
                            mapped_location = await self._resolve_existing_location_for_finalize(
                                db=db,
                                run=run,
                                item=parent_item,
                                company=company,
                            )
                            if mapped_location is not None:
                                parent_item.created_location_id = mapped_location.id
                                location_by_parent_item_id[parent_item.id] = mapped_location
                            else:
                                (
                                    location,
                                    created_new,
                                ) = await self._get_or_create_location_for_finalize(
                                    db=db,
                                    run=run,
                                    company_id=company.id,
                                    current_user=current_user,
                                    item=parent_item,
                                )
                                location_by_parent_item_id[parent_item.id] = location
                                if created_new:
                                    summary = BulkImportFinalizeSummary(
                                        run_id=run.id,
                                        locations_created=1,
                                        projects_created=0,
                                        rejected=0,
                                        invalid=0,
                                        duplicates_resolved=0,
                                    )
            if (
                treat_as_orphan
                and run.entrypoint_type == "company"
                and location_resolution is None
                and self._selected_existing_location_id(item) is None
                and self._orphan_project_create_new_location_data(item) is None
            ):
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="location_resolution required for orphan stream confirm",
                )
            (
                target_location,
                created_location_from_project_resolution,
            ) = await self._resolve_project_location_for_finalize(
                db=db,
                run=run,
                item=item,
                location_by_parent_item_id=location_by_parent_item_id,
                fallback_location=None,
                current_user=current_user,
                resolved_company=company,
                treat_as_orphan=treat_as_orphan,
            )
            locations_created = summary.locations_created + (
                1 if created_location_from_project_resolution else 0
            )
            project_data: dict[str, object] = {
                "technical_sections": copy.deepcopy(get_assessment_questionnaire()),
                WORKSPACE_PROJECT_DATA_KEY: WorkspaceService.build_workspace_v1_seed(
                    material_type=normalized.category,
                    material_name=normalized.name,
                    composition=normalized.description,
                    volume=normalized.volume or normalized.estimated_volume,
                    frequency=normalized.frequency,
                ),
            }
            if normalized.category and normalized.category.strip():
                project_data["bulk_import_category"] = normalized.category.strip()
            discovery_provenance = self._build_discovery_provenance_payload(
                run=run,
                discovery_source=discovery_source,
            )
            self._apply_workspace_provenance(
                project_data=project_data,
                provenance=discovery_provenance,
            )

            self._ensure_company_active_after_first_stream(company)
            project = Project(
                organization_id=run.organization_id,
                user_id=resolved_owner_user_id,
                location_id=target_location.id,
                name=normalized.name,
                client=company.name,
                sector=normalized.sector or company.sector,
                subsector=normalized.subsector or company.subsector,
                location=f"{target_location.name}, {target_location.city}",
                project_type=normalized.project_type,
                description=normalized.description,
                budget=0.0,
                schedule_summary="To be defined",
                tags=[],
                status="In Preparation",
                progress=0,
                project_data=project_data,
            )
            db.add(project)
            await db.flush()
            item.created_project_id = project.id
            registry = build_questionnaire_registry()
            self._create_pending_intake_suggestions_from_project_item(
                db=db,
                project=project,
                source_item=item,
                registry=registry,
                user_id=current_user.id,
            )
            summary = BulkImportFinalizeSummary(
                run_id=run.id,
                locations_created=locations_created,
                projects_created=1,
                rejected=0,
                invalid=0,
                duplicates_resolved=(
                    1 if item.duplicate_candidates and item.confirm_create_new else 0
                ),
            )
        else:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid action")

        await db.flush()
        pending_project_count_result = await db.execute(
            select(func.count(ImportItem.id)).where(
                ImportItem.run_id == run.id,
                ImportItem.organization_id == organization_id,
                ImportItem.item_type == "project",
                ImportItem.created_project_id.is_(None),
                ImportItem.status.in_(("pending_review", "accepted", "amended")),
            )
        )
        pending_project_count = int(pending_project_count_result.scalar_one() or 0)

        if run.source_type == "voice_interview":
            voice = await self._lock_voice_for_run(db, run_id=run.id)
            voice.status = "finalized" if pending_project_count == 0 else "partial_finalized"
            sync_import_run_status_for_voice(run=run, voice_status=voice.status)
            run.finalized_by_user_id = current_user.id
            if voice.status == "finalized":
                run.finalized_at = datetime.now(UTC)
        else:
            run.status = "completed" if pending_project_count == 0 else "review_ready"
            run.finalized_by_user_id = current_user.id
            if pending_project_count == 0:
                run.finalized_at = datetime.now(UTC)

        previous = self._summary_from_run(run) if run.summary_data else None
        accumulated = (
            BulkImportFinalizeSummary(
                run_id=run.id,
                locations_created=summary.locations_created,
                projects_created=summary.projects_created,
                rejected=summary.rejected,
                invalid=summary.invalid,
                duplicates_resolved=summary.duplicates_resolved,
            )
            if previous is None
            else BulkImportFinalizeSummary(
                run_id=run.id,
                locations_created=previous.locations_created + summary.locations_created,
                projects_created=previous.projects_created + summary.projects_created,
                rejected=previous.rejected + summary.rejected,
                invalid=previous.invalid + summary.invalid,
                duplicates_resolved=previous.duplicates_resolved + summary.duplicates_resolved,
            )
        )
        run.summary_data = accumulated.model_dump(mode="json")
        await self.refresh_run_counters(db, run)
        await db.flush()
        return item, accumulated, run

    async def _resolve_discovery_owner_user_id(
        self,
        db: AsyncSession,
        *,
        run: ImportRun,
        current_user: User,
        owner_user_id: UUID | None,
    ) -> UUID:
        default_owner_user_id = await self._resolve_run_owner_user_id(
            db,
            run=run,
            fallback_user_id=current_user.id,
        )
        if owner_user_id is None:
            return default_owner_user_id

        can_assign_owner = bool(current_user.is_superuser or current_user.role == "org_admin")
        if not can_assign_owner:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only org admins can assign owner",
            )

        owner = await db.get(User, owner_user_id)
        if owner is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Owner not found")
        if owner.organization_id != run.organization_id:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Owner must belong to your organization",
            )
        if not owner.is_active:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Owner must be active",
            )
        if owner.role not in {"org_admin", "field_agent"}:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Owner role is not allowed",
            )
        return owner.id

    async def _apply_company_resolution(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        item: ImportItem,
        company_resolution: BulkImportCompanyResolution,
    ) -> None:
        if run.entrypoint_type != "organization":
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="company_resolution only allowed for organization entrypoint",
            )

        existing_user_amendments = (
            item.user_amendments if isinstance(item.user_amendments, dict) else {}
        )
        merged_user_amendments = dict(existing_user_amendments)

        if isinstance(company_resolution, BulkImportCompanyResolutionExisting):
            selected_company = await db.get(Company, company_resolution.company_id)
            if (
                selected_company is None
                or selected_company.organization_id != run.organization_id
                or selected_company.archived_at is not None
            ):
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Selected company is invalid for this run",
                )
            merged_user_amendments["company_resolution"] = {
                "mode": "existing",
                "company_id": str(selected_company.id),
                "name": selected_company.name,
            }
            item.user_amendments = merged_user_amendments
            return

        if isinstance(company_resolution, BulkImportCompanyResolutionCreateNew):
            normalized_name = _sanitize_text(company_resolution.name, max_len=255)
            if not normalized_name:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="Company name is required",
                )
            merged_user_amendments["company_resolution"] = {
                "mode": "create_new",
                "name": normalized_name,
                "industry": _sanitize_text(company_resolution.industry, max_len=100),
                "sector": _sanitize_text(company_resolution.sector, max_len=50),
                "subsector": _sanitize_text(company_resolution.subsector, max_len=100),
            }
            item.user_amendments = merged_user_amendments
            return

        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Unsupported company_resolution mode",
        )

    def _selected_existing_company_id(self, item: ImportItem) -> UUID | None:
        amendments = item.user_amendments
        if not isinstance(amendments, dict):
            return None
        resolution = amendments.get("company_resolution")
        if not isinstance(resolution, dict):
            return None
        mode_raw = resolution.get("mode")
        company_id_raw = resolution.get("company_id")
        if mode_raw != "existing" or not isinstance(company_id_raw, str):
            return None
        try:
            return UUID(company_id_raw)
        except ValueError:
            return None

    def _create_new_company_payload(self, item: ImportItem) -> dict[str, str | None] | None:
        amendments = item.user_amendments
        if not isinstance(amendments, dict):
            return None
        resolution = amendments.get("company_resolution")
        if not isinstance(resolution, dict):
            return None
        mode_raw = resolution.get("mode")
        name_raw = resolution.get("name")
        if mode_raw != "create_new" or not isinstance(name_raw, str):
            return None

        normalized_name = _sanitize_text(name_raw, max_len=255)
        if not normalized_name:
            return None
        industry = resolution.get("industry")
        sector = resolution.get("sector")
        subsector = resolution.get("subsector")
        return {
            "name": normalized_name,
            "industry": _sanitize_text(
                industry if isinstance(industry, str) else None, max_len=100
            ),
            "sector": _sanitize_text(sector if isinstance(sector, str) else None, max_len=50),
            "subsector": _sanitize_text(
                subsector if isinstance(subsector, str) else None,
                max_len=100,
            ),
        }

    async def _resolve_company_for_discovery_draft(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        item: ImportItem,
        current_user: User,
    ) -> Company | None:
        if run.entrypoint_type == "company":
            return await self._load_entrypoint_company(db, run)
        if run.entrypoint_type != "organization":
            return None

        selected_company_id = self._selected_existing_company_id(item)
        if selected_company_id is not None:
            selected_company = await db.get(Company, selected_company_id)
            if (
                selected_company is None
                or selected_company.organization_id != run.organization_id
                or selected_company.archived_at is not None
            ):
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"Selected company invalid for item {item.id}",
                )
            return selected_company

        create_new_payload = self._create_new_company_payload(item)
        if create_new_payload is not None:
            existing_company = await self._find_company_by_name(
                db=db,
                organization_id=run.organization_id,
                name=create_new_payload["name"] or "",
            )
            if existing_company is not None:
                return existing_company

            company = Company(
                organization_id=run.organization_id,
                name=create_new_payload["name"] or "",
                industry=create_new_payload["industry"] or "Unknown",
                sector=create_new_payload["sector"] or "other",
                subsector=create_new_payload["subsector"] or "other",
                notes=None,
                tags=[],
                created_by_user_id=current_user.id,
            )
            db.add(company)
            await db.flush()
            return company

        company_name = _sanitize_text(
            str(item.normalized_data.get("company_name") or ""), max_len=255
        )
        if not company_name:
            company_name = _sanitize_text(
                str(item.normalized_data.get("client") or ""), max_len=255
            )
        if not company_name:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="company_resolution required for organization-scoped draft confirm",
            )

        matched_company = await self._find_company_by_name(
            db=db,
            organization_id=run.organization_id,
            name=company_name,
        )
        if matched_company is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="company_resolution required for organization-scoped draft confirm",
            )
        return matched_company

    async def _find_company_by_name(
        self,
        *,
        db: AsyncSession,
        organization_id: UUID,
        name: str,
    ) -> Company | None:
        normalized_name = _sanitize_text(name, max_len=255)
        if not normalized_name:
            return None
        result = await db.execute(
            select(Company)
            .where(
                Company.organization_id == organization_id,
                Company.archived_at.is_(None),
                func.lower(Company.name) == normalized_name.casefold(),
            )
            .limit(1)
        )
        return result.scalar_one_or_none()

    async def update_item_decision(
        self,
        db: AsyncSession,
        *,
        organization_id: UUID,
        item_id: UUID,
        action: str,
        normalized_data: dict[str, object] | None,
        review_notes: str | None,
        location_resolution: BulkImportLocationResolution | None,
        confirm_create_new: bool | None,
    ) -> ImportItem:
        run_result = await db.execute(
            select(ImportRun)
            .join(ImportItem, ImportItem.run_id == ImportRun.id)
            .where(ImportItem.id == item_id, ImportItem.organization_id == organization_id)
            .where(ImportRun.organization_id == organization_id)
            .with_for_update()
        )
        run = run_result.scalar_one_or_none()
        if not run:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")

        result = await db.execute(
            select(ImportItem)
            .where(ImportItem.id == item_id, ImportItem.organization_id == organization_id)
            .where(ImportItem.run_id == run.id)
            .with_for_update()
        )
        item = result.scalar_one_or_none()
        if not item:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
        if run.status != "review_ready":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Run is not editable",
            )

        if item.status == "invalid":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Invalid items are terminal",
            )

        if review_notes is not None:
            sanitized_review_notes = _sanitize_text(review_notes, max_len=1000)
            item.review_notes = sanitized_review_notes or None

        if location_resolution is not None and action != "amend":
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="location_resolution is only supported for amend",
            )

        if location_resolution is not None:
            await self._apply_location_resolution(
                db=db,
                run=run,
                item=item,
                location_resolution=location_resolution,
            )

        if confirm_create_new is not None:
            item.confirm_create_new = confirm_create_new

        if action == "accept":
            if not (
                run.source_type == "voice_interview"
                and item.duplicate_candidates
                and not item.confirm_create_new
            ):
                self._ensure_duplicate_confirmation(item, target_status="accepted")
            item.status = "accepted"
            item.needs_review = self._needs_review(item.item_type, item.normalized_data)
        elif action == "amend":
            if normalized_data is None and location_resolution is None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="normalized_data or location_resolution required for amend",
                )
            if normalized_data is not None:
                sanitized_patch = _sanitize_payload(normalized_data)
                merged = dict(item.normalized_data)
                merged.update(sanitized_patch)
                item.normalized_data = merged
                existing_user_amendments = (
                    item.user_amendments if isinstance(item.user_amendments, dict) else {}
                )
                merged_user_amendments = dict(existing_user_amendments)
                merged_user_amendments.update(sanitized_patch)
                item.user_amendments = merged_user_amendments
            if not (
                run.source_type == "voice_interview"
                and item.duplicate_candidates
                and not item.confirm_create_new
            ):
                self._ensure_duplicate_confirmation(item, target_status="amended")
            item.status = "amended"
            item.needs_review = self._needs_review(item.item_type, item.normalized_data)
        elif action == "reject":
            item.status = "rejected"
            item.needs_review = False
            if item.item_type == "location":
                await db.execute(
                    update(ImportItem)
                    .where(ImportItem.parent_item_id == item.id)
                    .where(ImportItem.run_id == item.run_id)
                    .values(status="rejected", needs_review=False)
                )
        elif action == "reset":
            item.status = "pending_review"
            item.needs_review = True
            item.confirm_create_new = False
        else:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid action")

        await self.refresh_run_counters(db, run)
        await db.flush()
        return item

    async def get_effective_company_id_for_run(self, db: AsyncSession, *, run: ImportRun) -> UUID:
        if run.entrypoint_type == "company":
            company = await self._load_entrypoint_company(db, run)
            if company is None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Company entrypoint required",
                )
            return company.id

        entrypoint_location = await db.get(Location, run.entrypoint_id)
        if (
            entrypoint_location is None
            or entrypoint_location.organization_id != run.organization_id
        ):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Entrypoint location not found",
            )
        return entrypoint_location.company_id

    async def _apply_location_resolution(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        item: ImportItem,
        location_resolution: BulkImportLocationResolution,
        effective_company_id_override: UUID | None = None,
    ) -> None:
        if item.item_type == "project":
            if item.parent_item_id is not None:
                return
            await self._apply_project_location_resolution(
                db=db,
                run=run,
                item=item,
                location_resolution=location_resolution,
            )
            return
        if item.item_type != "location":
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="location_resolution only allowed for location/project items",
            )

        effective_company_id = (
            effective_company_id_override
            if effective_company_id_override is not None
            else await self.get_effective_company_id_for_run(db, run=run)
        )

        if isinstance(location_resolution, BulkImportLocationResolutionExisting):
            selected_location = await db.get(Location, location_resolution.location_id)
            if (
                selected_location is None
                or selected_location.organization_id != run.organization_id
                or selected_location.company_id != effective_company_id
            ):
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Selected location is invalid for this run",
                )

            item.created_location_id = None
            item.normalized_data = {
                "name": selected_location.name,
                "city": selected_location.city,
                "state": selected_location.state,
                "address": selected_location.address,
            }
            item.user_amendments = {
                "name": selected_location.name,
                "city": selected_location.city,
                "state": selected_location.state,
                "address": selected_location.address,
                "location_resolution": {
                    "mode": "existing",
                    "location_id": str(selected_location.id),
                },
            }
            item.confirm_create_new = False
            return

        if isinstance(location_resolution, BulkImportLocationResolutionCreateNew):
            normalized_location = NormalizedLocationDataV1(
                name=location_resolution.name,
                city=location_resolution.city,
                state=location_resolution.state,
                address=location_resolution.address,
            )
            normalized_payload = normalized_location.model_dump(mode="json")

            item.created_location_id = None
            item.normalized_data = normalized_payload
            item.user_amendments = {
                **normalized_payload,
                "location_resolution": {
                    "mode": "create_new",
                    "name": normalized_location.name,
                    "city": normalized_location.city,
                    "state": normalized_location.state,
                    "address": normalized_location.address,
                },
            }
            item.confirm_create_new = True
            return

        if isinstance(location_resolution, BulkImportLocationResolutionLocked):
            normalized_existing = self._validated_location_data(item)
            locked_name = _sanitize_text(location_resolution.name, max_len=255)
            if not locked_name:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="Locked location name is required",
                )

            item.created_location_id = None
            item.normalized_data = {
                "name": locked_name,
                "city": normalized_existing.city,
                "state": normalized_existing.state,
                "address": normalized_existing.address,
            }
            item.user_amendments = {
                "name": locked_name,
                "city": normalized_existing.city,
                "state": normalized_existing.state,
                "address": normalized_existing.address,
                "location_resolution": {
                    "mode": "locked",
                    "name": locked_name,
                },
            }
            return

        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Unsupported location_resolution mode",
        )

    async def _apply_project_location_resolution(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        item: ImportItem,
        location_resolution: BulkImportLocationResolution,
    ) -> None:
        if run.entrypoint_type not in {"company", "organization"}:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="project location_resolution requires company or organization entrypoint",
            )
        if item.parent_item_id is not None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="project location_resolution only allowed for orphan stream drafts",
            )

        effective_company_id: UUID | None
        if run.entrypoint_type == "company":
            effective_company_id = await self.get_effective_company_id_for_run(db, run=run)
        else:
            effective_company_id = self._selected_existing_company_id(item)
        existing_user_amendments = (
            item.user_amendments if isinstance(item.user_amendments, dict) else {}
        )
        merged_user_amendments = dict(existing_user_amendments)

        if isinstance(location_resolution, BulkImportLocationResolutionExisting):
            selected_location = await db.get(Location, location_resolution.location_id)
            if (
                selected_location is None
                or selected_location.organization_id != run.organization_id
                or (
                    effective_company_id is not None
                    and selected_location.company_id != effective_company_id
                )
            ):
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Selected location is invalid for this run",
                )

            merged_user_amendments["location_resolution"] = {
                "mode": "existing",
                "location_id": str(selected_location.id),
                "name": selected_location.name,
                "city": selected_location.city,
                "state": selected_location.state,
                "address": selected_location.address,
            }
            item.user_amendments = merged_user_amendments
            return

        if isinstance(location_resolution, BulkImportLocationResolutionCreateNew):
            normalized_location = NormalizedLocationDataV1(
                name=location_resolution.name,
                city=location_resolution.city,
                state=location_resolution.state,
                address=location_resolution.address,
            )
            merged_user_amendments["location_resolution"] = {
                "mode": "create_new",
                "name": normalized_location.name,
                "city": normalized_location.city,
                "state": normalized_location.state,
                "address": normalized_location.address,
            }
            item.user_amendments = merged_user_amendments
            return

        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Orphan stream location resolution must be existing or create_new",
        )

    async def import_orphan_projects(
        self,
        db: AsyncSession,
        *,
        organization_id: UUID,
        run_id: UUID,
        location_id: UUID,
        item_ids: list[UUID],
        user_id: UUID,
    ) -> dict[str, object]:
        """Create Projects directly from orphan items' normalizedData.

        No re-analysis — reads the already-extracted data and creates entities.
        """
        # Lock the run
        result = await db.execute(
            select(ImportRun)
            .where(
                ImportRun.id == run_id,
                ImportRun.organization_id == organization_id,
            )
            .with_for_update()
        )
        run = result.scalar_one_or_none()
        if not run:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Run not found")
        if run.status != "review_ready":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Run must be in review_ready status",
            )
        if run.entrypoint_type != "company":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Only company-entrypoint runs can have orphan projects",
            )
        if run.source_type == "voice_interview":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Voice orphan import is disabled. Resolve from Needs Confirmation.",
            )

        # Validate location
        location = await db.get(Location, location_id)
        if not location or location.organization_id != organization_id:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Location not found")
        if location.company_id != run.entrypoint_id:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Location does not belong to the run's company",
            )

        # Load company for project fields
        company = await db.get(Company, location.company_id)
        if not company:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Company not found")
        discovery_provenance = await self._build_discovery_provenance_for_run(db=db, run=run)

        # Load requested items (locked)
        items_result = await db.execute(
            select(ImportItem)
            .where(
                ImportItem.run_id == run.id,
                ImportItem.id.in_(item_ids),
            )
            .with_for_update()
        )
        items = items_result.scalars().all()

        # Ensure all requested IDs were found in this run
        if len(items) != len(item_ids):
            found_ids = {i.id for i in items}
            missing = [str(uid) for uid in item_ids if uid not in found_ids]
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Items not found in this run: {', '.join(missing)}",
            )

        # Process each item
        created_project_ids: dict[str, str] = {}
        skipped = 0
        orphan_eligible_statuses = {"invalid", "pending_review", "accepted", "amended"}
        for item in items:
            if item.item_type != "project":
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"Item {item.id} is not a project",
                )
            if item.status not in orphan_eligible_statuses or item.parent_item_id is not None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"Item {item.id} is not an orphan (status={item.status})",
                )
            if item.created_project_id is not None:
                skipped += 1
                continue  # already imported, skip

            normalized = self._validated_project_data(item)

            # Duplicate guard: same name in same location
            existing = await db.execute(
                select(Project.id).where(
                    Project.location_id == location.id,
                    func.lower(Project.name) == normalized.name.strip().lower(),
                )
            )
            if existing.scalars().first():
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"A project named '{normalized.name}' already exists in this location",
                )

            # Create project (same mapping as finalize_run)
            project_data: dict[str, object] = {
                "technical_sections": copy.deepcopy(get_assessment_questionnaire()),
                WORKSPACE_PROJECT_DATA_KEY: WorkspaceService.build_workspace_v1_seed(
                    material_type=normalized.category,
                    material_name=normalized.name,
                    composition=normalized.description,
                    volume=normalized.volume or normalized.estimated_volume,
                    frequency=normalized.frequency,
                ),
            }
            if normalized.category and normalized.category.strip():
                project_data["bulk_import_category"] = normalized.category.strip()
            self._apply_workspace_provenance(
                project_data=project_data,
                provenance=discovery_provenance,
            )

            project = Project(
                organization_id=organization_id,
                user_id=user_id,
                location_id=location.id,
                name=normalized.name,
                client=company.name,
                sector=normalized.sector or company.sector,
                subsector=normalized.subsector or company.subsector,
                location=f"{location.name}, {location.city}",
                project_type=normalized.project_type,
                description=normalized.description,
                budget=0.0,
                schedule_summary="To be defined",
                tags=[],
                status="In Preparation",
                progress=0,
                project_data=project_data,
            )
            db.add(project)
            await db.flush()
            item.created_project_id = project.id
            created_project_ids[str(item.id)] = str(project.id)

        await self.refresh_run_counters(db, run)

        # If all orphan project items now have created_project_id, mark run completed
        unresolved = await db.execute(
            select(func.count()).where(
                ImportItem.run_id == run.id,
                ImportItem.item_type == "project",
                ImportItem.parent_item_id.is_(None),
                ImportItem.status.in_(tuple(orphan_eligible_statuses)),
                ImportItem.created_project_id.is_(None),
            )
        )
        if unresolved.scalar_one() == 0:
            run.status = "completed"
            run.finalized_at = datetime.now(UTC)
            await db.flush()

        return {
            "projects_created": len(created_project_ids),
            "created_project_ids": created_project_ids,
            "skipped": skipped,
        }

    async def get_run(
        self, db: AsyncSession, *, organization_id: UUID, run_id: UUID
    ) -> ImportRun | None:
        result = await db.execute(
            select(ImportRun)
            .where(ImportRun.id == run_id, ImportRun.organization_id == organization_id)
            .limit(1)
        )
        return result.scalar_one_or_none()

    async def get_item(
        self, db: AsyncSession, *, organization_id: UUID, item_id: UUID
    ) -> ImportItem | None:
        result = await db.execute(
            select(ImportItem)
            .where(ImportItem.id == item_id, ImportItem.organization_id == organization_id)
            .limit(1)
        )
        return result.scalar_one_or_none()

    async def list_items(
        self,
        db: AsyncSession,
        *,
        organization_id: UUID,
        run_id: UUID,
        page: int,
        size: int,
        status_filter: str | None,
    ) -> tuple[list[ImportItem], int]:
        query = select(ImportItem).where(
            ImportItem.run_id == run_id,
            ImportItem.organization_id == organization_id,
        )
        if status_filter:
            query = query.where(ImportItem.status == status_filter)

        count_result = await db.execute(select(func.count()).select_from(query.subquery()))
        total = int(count_result.scalar_one() or 0)

        paged_result = await db.execute(
            query.order_by(ImportItem.created_at, ImportItem.id)
            .offset((page - 1) * size)
            .limit(size)
        )
        return list(paged_result.scalars().all()), total

    def _summary_from_run(self, run: ImportRun) -> BulkImportFinalizeSummary:
        if not run.summary_data:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Run completed without summary",
            )
        return BulkImportFinalizeSummary.model_validate(run.summary_data)

    def get_run_summary(self, run: ImportRun) -> BulkImportFinalizeSummary:
        return self._summary_from_run(run)

    async def _build_discovery_provenance_for_run(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
    ) -> dict[str, object] | None:
        discovery_source_result = await db.execute(
            select(DiscoverySource).where(DiscoverySource.import_run_id == run.id).limit(1)
        )
        discovery_source = discovery_source_result.scalar_one_or_none()
        return self._build_discovery_provenance_payload(
            run=run,
            discovery_source=discovery_source,
        )

    def _build_discovery_provenance_payload(
        self,
        *,
        run: ImportRun,
        discovery_source: DiscoverySource | None,
    ) -> dict[str, object] | None:
        if discovery_source is None:
            return None
        if discovery_source.source_type not in {"file", "audio", "text"}:
            return None
        return {
            "origin": "ai_discovery",
            "run_id": str(run.id),
            "discovery_session_id": (
                str(discovery_source.session_id) if discovery_source.session_id else None
            ),
            "source_type": discovery_source.source_type,
            "source_filename": discovery_source.source_filename,
            "discovery_source_id": (str(discovery_source.id) if discovery_source.id else None),
        }

    def _apply_workspace_provenance(
        self,
        *,
        project_data: dict[str, object],
        provenance: dict[str, object] | None,
    ) -> None:
        if provenance is None:
            return
        workspace_raw = project_data.get(WORKSPACE_PROJECT_DATA_KEY)
        workspace_data = workspace_raw if isinstance(workspace_raw, dict) else {}
        merged_workspace_data = dict(workspace_data)
        merged_workspace_data["provenance"] = provenance
        project_data[WORKSPACE_PROJECT_DATA_KEY] = merged_workspace_data

    def _legacy_orphan_group_id(self, *, run_id: UUID, item_id: UUID) -> str:
        return f"legacy_orphan:{run_id}:{item_id}"

    def _legacy_linked_group_id(self, *, run_id: UUID, anchor_item_id: UUID) -> str:
        return f"legacy_linked:{run_id}:{anchor_item_id}"

    async def _assert_finalize_ready(self, db: AsyncSession, run: ImportRun) -> None:
        pending_count_result = await db.execute(
            select(func.count(ImportItem.id)).where(
                ImportItem.run_id == run.id,
                ImportItem.status == "pending_review",
            )
        )
        pending_count = int(pending_count_result.scalar_one() or 0)
        if pending_count > 0:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Run has pending_review items",
            )

        items_result = await db.execute(select(ImportItem).where(ImportItem.run_id == run.id))
        items = items_result.scalars().all()
        by_id = {item.id: item for item in items}

        for item in items:
            if item.status not in {"accepted", "amended"}:
                continue
            self._ensure_duplicate_confirmation(item)
            if item.needs_review:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"Item {item.id} still needs review",
                )
            if item.item_type == "location":
                self._validated_location_data(item)
                if run.entrypoint_type == "location":
                    raise HTTPException(
                        status_code=status.HTTP_409_CONFLICT,
                        detail="Location items not allowed for location entrypoint",
                    )
            else:
                self._validated_project_data(item)
                if item.parent_item_id:
                    parent = by_id.get(item.parent_item_id)
                    if parent is None or parent.item_type != "location":
                        raise HTTPException(
                            status_code=status.HTTP_409_CONFLICT,
                            detail="Project item parent invalid",
                        )
                    if parent.status not in {"accepted", "amended"}:
                        raise HTTPException(
                            status_code=status.HTTP_409_CONFLICT,
                            detail="Project item parent is not accepted",
                        )

    def _effective_group_id(self, item: ImportItem) -> str | None:
        if item.group_id is not None:
            return item.group_id
        if item.item_type == "project" and item.parent_item_id is not None:
            return self._legacy_linked_group_id(
                run_id=item.run_id,
                anchor_item_id=item.parent_item_id,
            )
        if item.item_type == "project" and item.parent_item_id is None:
            return self._legacy_orphan_group_id(run_id=item.run_id, item_id=item.id)
        return None

    def _effective_group_id_with_index(
        self,
        item: ImportItem,
        *,
        by_id: dict[UUID, ImportItem],
    ) -> str | None:
        effective_group_id = self._effective_group_id(item)
        if effective_group_id is not None:
            return effective_group_id
        if item.item_type == "location":
            legacy_child = next(
                (
                    child
                    for child in by_id.values()
                    if child.parent_item_id == item.id and child.item_type == "project"
                ),
                None,
            )
            if legacy_child is not None:
                return self._legacy_linked_group_id(
                    run_id=item.run_id,
                    anchor_item_id=item.id,
                )
        return None

    async def _resolve_project_location_for_finalize(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        item: ImportItem,
        location_by_parent_item_id: dict[UUID, Location],
        fallback_location: Location | None,
        current_user: User,
        resolved_company: Company | None = None,
        treat_as_orphan: bool = False,
    ) -> tuple[Location, bool]:
        if run.entrypoint_type == "location":
            if fallback_location is None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Entrypoint location missing",
                )
            return fallback_location, False

        if (
            not treat_as_orphan
            and item.parent_item_id
            and item.parent_item_id in location_by_parent_item_id
        ):
            return location_by_parent_item_id[item.parent_item_id], False

        (
            project_resolution_location,
            created_new_from_project_resolution,
        ) = await self._resolve_orphan_project_location_for_finalize(
            db=db,
            run=run,
            item=item,
            current_user=current_user,
            resolved_company=resolved_company,
            treat_as_orphan=treat_as_orphan,
        )
        if project_resolution_location is not None:
            return project_resolution_location, created_new_from_project_resolution

        location_name = _sanitize_text(str(item.normalized_data.get("location_name") or ""))
        city = _sanitize_text(str(item.normalized_data.get("location_city") or ""))
        state_value = _sanitize_text(str(item.normalized_data.get("location_state") or ""))
        if not location_name or not city or not state_value:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Project item location unresolved",
            )

        company = await self._load_entrypoint_company(db, run)
        if company is None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Company entrypoint required",
            )

        result = await db.execute(
            select(Location).where(
                Location.organization_id == run.organization_id,
                Location.company_id == company.id,
                func.lower(Location.name) == location_name.casefold(),
                func.lower(Location.city) == city.casefold(),
                func.lower(Location.state) == state_value.casefold(),
            )
        )
        location = result.scalar_one_or_none()
        if not location:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Project item location not found",
            )
        return location, False

    async def _resolve_orphan_project_location_for_finalize(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        item: ImportItem,
        current_user: User,
        resolved_company: Company | None = None,
        treat_as_orphan: bool = False,
    ) -> tuple[Location | None, bool]:
        if run.entrypoint_type not in {"company", "organization"}:
            return None, False
        if item.parent_item_id is not None and not treat_as_orphan:
            return None, False

        company = resolved_company
        if company is None and run.entrypoint_type == "company":
            company = await self._load_entrypoint_company(db, run)
        if company is None:
            return None, False

        selected_existing_location_id = self._selected_existing_location_id(item)
        if selected_existing_location_id is not None:
            selected_location = await db.get(Location, selected_existing_location_id)
            if (
                selected_location is None
                or selected_location.organization_id != run.organization_id
                or selected_location.company_id != company.id
            ):
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"Selected location invalid for item {item.id}",
                )
            return selected_location, False

        create_new_location_data = self._orphan_project_create_new_location_data(item)
        if create_new_location_data is None:
            return None, False

        existing_location = await self._find_location_by_identity(
            db=db,
            organization_id=run.organization_id,
            company_id=company.id,
            location_data=create_new_location_data,
        )
        if existing_location is not None:
            return existing_location, False

        location = Location(
            organization_id=run.organization_id,
            company_id=company.id,
            name=create_new_location_data.name,
            city=create_new_location_data.city,
            state=create_new_location_data.state,
            address=create_new_location_data.address,
            created_by_user_id=current_user.id,
        )
        db.add(location)
        await db.flush()
        return location, True

    def _orphan_project_create_new_location_data(
        self,
        item: ImportItem,
    ) -> NormalizedLocationDataV1 | None:
        amendments = item.user_amendments
        if not isinstance(amendments, dict):
            return None
        resolution = amendments.get("location_resolution")
        if not isinstance(resolution, dict):
            return None

        mode_raw: object | None = None
        name_raw: object | None = None
        city_raw: object | None = None
        state_raw: object | None = None
        address_raw: object | None = None

        for key, value in resolution.items():
            if key == "mode":
                mode_raw = value
            elif key == "name":
                name_raw = value
            elif key == "city":
                city_raw = value
            elif key == "state":
                state_raw = value
            elif key == "address":
                address_raw = value

        if mode_raw != "create_new":
            return None
        if (
            not isinstance(name_raw, str)
            or not isinstance(city_raw, str)
            or not isinstance(state_raw, str)
        ):
            return None
        address_value = address_raw if isinstance(address_raw, str) else None

        try:
            return NormalizedLocationDataV1(
                name=name_raw,
                city=city_raw,
                state=state_raw,
                address=address_value,
            )
        except Exception:
            return None

    def _ensure_duplicate_confirmation(
        self, item: ImportItem, target_status: str | None = None
    ) -> None:
        status_to_validate = target_status or item.status
        if status_to_validate not in {"accepted", "amended"}:
            return
        candidates = item.duplicate_candidates or []
        if len(candidates) > 0 and not item.confirm_create_new:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Duplicate requires confirm_create_new=true",
            )

    def _validated_location_data(self, item: ImportItem) -> NormalizedLocationDataV1:
        try:
            return NormalizedLocationDataV1.model_validate(item.normalized_data)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Location item invalid: {item.id}",
            ) from exc

    def _validated_project_data(self, item: ImportItem) -> NormalizedProjectDataV1:
        try:
            data = NormalizedProjectDataV1.model_validate(item.normalized_data)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Project item invalid: {item.id}",
            ) from exc
        return data

    async def _load_entrypoint_company(self, db: AsyncSession, run: ImportRun) -> Company | None:
        if run.entrypoint_type != "company":
            return None
        company = await db.get(Company, run.entrypoint_id)
        if not company or company.organization_id != run.organization_id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Entrypoint company not found"
            )
        return company

    def _needs_review(self, item_type: str, payload: dict[str, object]) -> bool:
        if item_type == "location":
            fields = [
                _sanitize_text(str(payload.get("name") or "")),
                _sanitize_text(str(payload.get("city") or "")),
                _sanitize_text(str(payload.get("state") or "")),
            ]
            return any(not field for field in fields)
        fields = [_sanitize_text(str(payload.get("name") or ""))]
        return any(not field for field in fields)

    def _needs_review_by_confidence(self, confidence: int) -> bool:
        return confidence < 80

    def _confidence_from_raw(self, raw: dict[str, str], key: str, default: int) -> int:
        value = raw.get(key)
        if not value:
            return default
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            return default
        if parsed < 0:
            return 0
        if parsed > 100:
            return 100
        return parsed

    async def _handle_processing_failure(
        self,
        db: AsyncSession,
        run_id: UUID,
        exc: Exception,
    ) -> None:
        retryable = not isinstance(exc, (ParserLimitError, ValueError))
        reason = _truncate_error(str(exc) or "processing_failed")

        run_result = await db.execute(
            select(ImportRun).where(ImportRun.id == run_id).with_for_update()
        )
        run = run_result.scalar_one_or_none()
        if run is None:
            return

        voice: VoiceInterview | None = None
        if run.source_type == "voice_interview":
            voice_result = await db.execute(
                select(VoiceInterview)
                .where(VoiceInterview.bulk_import_run_id == run_id)
                .with_for_update()
            )
            voice = voice_result.scalar_one_or_none()

        failed_stage = None
        reason_code = reason.casefold()
        if "transcrib" in reason_code:
            failed_stage = "transcribing"
        elif "extract" in reason_code or "voice_transcript" in reason_code:
            failed_stage = "extracting"

        if retryable and run.processing_attempts < MAX_PROCESSING_ATTEMPTS:
            if run.source_type == "voice_interview":
                if voice is None:
                    raise ValueError("voice_interview_not_found")
                voice.status = "queued"
                voice.error_code = reason
                voice.failed_stage = failed_stage
                voice.processing_attempts = run.processing_attempts
                sync_import_run_status_for_voice(run=run, voice_status=voice.status)
            else:
                run.status = "uploaded"
            run.progress_step = None
            run.processing_error = reason
            run.processing_available_at = datetime.now(UTC) + timedelta(
                seconds=_dedupe_backoff_seconds(run.id, run.processing_attempts)
            )
            run.processing_started_at = None
            await self.sync_discovery_session_for_run(db, run_id=run.id)
            await db.flush()
            return

        if run.source_type == "voice_interview":
            if voice is None:
                raise ValueError("voice_interview_not_found")
            voice.status = "failed"
            voice.error_code = reason
            voice.failed_stage = failed_stage
            voice.processing_attempts = run.processing_attempts
            sync_import_run_status_for_voice(run=run, voice_status=voice.status)
        else:
            run.status = "failed"
        run.progress_step = None
        run.processing_error = reason
        await self.sync_discovery_session_for_run(db, run_id=run.id)
        await db.flush()

    async def _parse_source_with_hard_timeout(
        self,
        *,
        filename: str,
        file_bytes: bytes,
        timeout_seconds: float,
        parse_callable: Callable[[str, bytes], list[ParsedRow]] | None = None,
    ) -> list[ParsedRow]:
        """Run parser in killable subprocess; enforce real timeout cancelation."""
        parser = parse_callable or _default_parse_callable
        ctx = multiprocessing.get_context("spawn")
        parent_conn, child_conn = ctx.Pipe(duplex=False)
        process = ctx.Process(
            name=PARSER_WORKER_NAME,
            target=_parse_source_subprocess_entrypoint,
            args=(parser, filename, file_bytes, child_conn),
            daemon=True,
        )

        process.start()
        child_conn.close()
        payload: tuple[object, ...] | None = None
        timed_out = False
        try:
            deadline = time.monotonic() + timeout_seconds
            while True:
                remaining = deadline - time.monotonic()
                if remaining <= 0:
                    timed_out = True
                    break

                if parent_conn.poll(min(0.1, remaining)):
                    try:
                        payload = parent_conn.recv()
                    except EOFError:
                        payload = None
                    break

                if process.exitcode is not None and not parent_conn.poll():
                    break

            if payload is None:
                if process.is_alive():
                    process.terminate()
                    await asyncio.to_thread(process.join, 1.0)
                    if process.is_alive():
                        process.kill()
                        await asyncio.to_thread(process.join, 1.0)
                if timed_out:
                    raise ValueError("parser_timeout")
                raise ValueError("parser_no_result")

            await asyncio.to_thread(process.join, 1.0)
            if process.is_alive():
                process.terminate()
                await asyncio.to_thread(process.join, 1.0)
                if process.is_alive():
                    process.kill()
                    await asyncio.to_thread(process.join, 1.0)
        finally:
            parent_conn.close()

        status_code = payload[0]
        if status_code == "ok_file":
            temp_path = payload[1]
            if not isinstance(temp_path, str):
                raise ValueError("parser_invalid_result_path")
            temp_file_path = Path(temp_path)
            try:
                serialized = await asyncio.to_thread(_load_json_parse_result, temp_file_path)
            finally:
                try:
                    await asyncio.to_thread(temp_file_path.unlink, missing_ok=True)
                except OSError:
                    logger.warning("bulk_import_parse_tempfile_cleanup_failed", path=temp_path)
            return _deserialize_parsed_rows(serialized)

        if status_code == "error":
            error_type = payload[1]
            error_message = payload[2]
            if error_type == "ParserLimitError":
                raise ParserLimitError(error_message)
            if error_type == "ValueError":
                raise ValueError(error_message)
            raise RuntimeError(error_message)

        raise ValueError("parser_invalid_payload")

    def _parse_source(self, filename: str, file_bytes: bytes) -> list[ParsedRow]:
        extension = Path(filename).suffix.casefold()
        if extension == ".csv":
            rows = self._parse_csv(file_bytes)
        elif extension == ".xlsx":
            rows = self._parse_excel(file_bytes)
        elif extension == ".xls":
            raise ValueError("legacy_xls_not_supported")
        elif extension == ".pdf":
            rows = self._parse_pdf(file_bytes)
        else:
            raise ValueError("unsupported_file_type")

        parsed: list[ParsedRow] = []
        for row in rows:
            normalized_row = {
                str(k).strip().casefold(): _sanitize_text(str(v) if v is not None else "")
                for k, v in row.items()
            }
            location_data = self._extract_location_data(normalized_row)
            project_data = self._extract_project_data(normalized_row)
            if not location_data and not project_data:
                continue
            parsed.append(
                ParsedRow(
                    location_data=location_data,
                    project_data=project_data,
                    raw={k: v or "" for k, v in normalized_row.items()},
                )
            )
        return parsed

    def _parse_csv(self, file_bytes: bytes) -> list[dict[str, str]]:
        text_value = file_bytes.decode("utf-8-sig", errors="ignore")
        return self._parse_delimited_text(text_value)

    def _parse_excel(self, file_bytes: bytes) -> list[dict[str, str]]:
        load_workbook = self._get_openpyxl_load_workbook()
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        try:
            sheets = workbook.worksheets
            if not sheets:
                return []
            sheet = sheets[0]
            rows_iter = sheet.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return []
            headers = [str(value).strip() if value is not None else "" for value in header_row]

            rows: list[dict[str, str]] = []
            row_count = 0
            cell_count = 0
            for values in rows_iter:
                row_count += 1
                if row_count > MAX_IMPORT_ROWS:
                    raise ParserLimitError("max_rows_exceeded")
                cell_count += len(values)
                if cell_count > MAX_IMPORT_CELLS:
                    raise ParserLimitError("max_cells_exceeded")
                row: dict[str, str] = {}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    value = values[index] if index < len(values) else None
                    row[header] = "" if value is None else str(value)
                rows.append(row)
            return rows
        finally:
            workbook.close()

    def _get_openpyxl_load_workbook(self):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("xlsx_parser_unavailable") from exc
        return load_workbook

    def _assert_xlsx_parser_available(self) -> None:
        self._get_openpyxl_load_workbook()

    def _parse_pdf(self, file_bytes: bytes) -> list[dict[str, str]]:
        reader = PdfReader(io.BytesIO(file_bytes))
        lines: list[str] = []
        row_count = 0
        for page in reader.pages:
            text_value = page.extract_text() or ""
            page_lines = [line.strip() for line in text_value.splitlines() if line.strip()]
            lines.extend(page_lines)
            row_count += len(page_lines)
            if row_count > MAX_IMPORT_ROWS:
                raise ParserLimitError("max_rows_exceeded")
        return self._parse_text_lines(lines)

    def _parse_delimited_text(self, text_value: str) -> list[dict[str, str]]:
        sample = text_value[:1024]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.get_dialect("excel")

        reader = csv.DictReader(io.StringIO(text_value), dialect=dialect)
        rows: list[dict[str, str]] = []
        cell_count = 0
        for row_count, row in enumerate(reader, start=1):
            if row_count > MAX_IMPORT_ROWS:
                raise ParserLimitError("max_rows_exceeded")
            cell_count += len(row)
            if cell_count > MAX_IMPORT_CELLS:
                raise ParserLimitError("max_cells_exceeded")
            rows.append({str(k): "" if v is None else str(v) for k, v in row.items()})
        return rows

    def _parse_text_lines(self, lines: list[str]) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        current_location_name = ""
        current_city = ""
        current_state = ""
        current_category = ""

        for line in lines:
            if ":" in line:
                prefix, value = line.split(":", 1)
                key = prefix.strip().casefold()
                content = value.strip()
                if key in {"location", "site", "plant", "location name"}:
                    current_location_name = content
                    continue
                if key == "city":
                    current_city = content
                    continue
                if key == "state":
                    current_state = content
                    continue
                if key == "category":
                    current_category = content
                    continue
                if key in {"project", "waste stream", "stream"}:
                    rows.append(
                        {
                            "location_name": current_location_name,
                            "city": current_city,
                            "state": current_state,
                            "project_name": content,
                            "category": current_category,
                        }
                    )
                    if len(rows) > MAX_IMPORT_ROWS:
                        raise ParserLimitError("max_rows_exceeded")
                    continue

            comma_parts = [part.strip() for part in line.split(",")]
            if len(comma_parts) >= 2:
                row: dict[str, str] = {
                    "project_name": comma_parts[0],
                    "category": comma_parts[1],
                }
                if len(comma_parts) >= 5:
                    row["location_name"] = comma_parts[2]
                    row["city"] = comma_parts[3]
                    row["state"] = comma_parts[4]
                rows.append(row)
                if len(rows) > MAX_IMPORT_ROWS:
                    raise ParserLimitError("max_rows_exceeded")

        return rows

    def _extract_location_data(self, row: dict[str, str | None]) -> dict[str, str] | None:
        name = self._pick_value(
            row, ["location_name", "location", "site", "plant", "location name"]
        )
        city = self._pick_value(row, ["city", "location_city"])
        state_value = self._pick_value(row, ["state", "province", "location_state"])
        address = self._pick_value(row, ["address", "location_address"])
        if not any([name, city, state_value, address]):
            return None
        return {
            "name": name,
            "city": city,
            "state": state_value,
            "address": address,
        }

    def _extract_project_data(self, row: dict[str, str | None]) -> dict[str, str] | None:
        name = self._pick_value(
            row, ["project_name", "waste_stream", "waste stream", "project", "name"]
        )
        category = self._pick_value(row, ["category", "waste_category", "waste category"])
        project_type = self._pick_value(row, ["project_type", "project type"]) or "Assessment"
        description = self._pick_value(row, ["description", "details"])
        sector = self._pick_value(row, ["sector"])
        subsector = self._pick_value(row, ["subsector", "sub_sector"])
        estimated_volume = self._pick_value(row, ["estimated_volume", "volume", "estimated volume"])
        volume = self._pick_value(row, ["volume"])
        units = self._pick_value(row, ["units", "unit"])
        frequency = self._pick_value(row, ["frequency"])

        if not any([name, category, description, estimated_volume]):
            return None

        return {
            "name": name,
            "category": category,
            "project_type": project_type,
            "description": description,
            "sector": sector,
            "subsector": subsector,
            "estimated_volume": estimated_volume,
            "volume": volume,
            "units": units,
            "frequency": frequency,
        }

    def _pick_value(self, row: dict[str, str | None], keys: list[str]) -> str:
        for key in keys:
            value = row.get(key)
            if value and value.strip():
                return value.strip()
        return ""

    async def _build_import_items(
        self,
        db: AsyncSession,
        run: ImportRun,
        parsed_rows: list[ParsedRow],
    ) -> list[ImportItem]:
        if not parsed_rows:
            return []

        if run.entrypoint_type == "organization":
            return await self._build_organization_entrypoint_items(db, run, parsed_rows)

        if run.entrypoint_type == "company":
            return await self._build_company_entrypoint_items(db, run, parsed_rows)
        return await self._build_location_entrypoint_items(db, run, parsed_rows)

    async def build_items_for_parsed_rows(
        self,
        db: AsyncSession,
        *,
        run: ImportRun,
        parsed_rows: list[ParsedRow],
    ) -> list[ImportItem]:
        """Public wrapper for building staged items from parsed rows."""

        return await self._build_import_items(db, run, parsed_rows)

    async def _build_company_entrypoint_items(
        self,
        db: AsyncSession,
        run: ImportRun,
        parsed_rows: list[ParsedRow],
    ) -> list[ImportItem]:
        company = await db.get(Company, run.entrypoint_id)
        if not company or company.organization_id != run.organization_id:
            raise ValueError("entrypoint_company_not_found")

        existing_locations_result = await db.execute(
            select(Location).where(
                Location.organization_id == run.organization_id,
                Location.company_id == company.id,
            )
        )
        existing_locations = existing_locations_result.scalars().all()
        existing_location_ids_by_key: dict[str, list[UUID]] = {}
        for existing in existing_locations:
            location_key = self._location_key(
                {
                    "name": existing.name,
                    "city": existing.city,
                    "state": existing.state,
                }
            )
            existing_location_ids_by_key.setdefault(location_key, []).append(existing.id)

        location_defs: dict[str, dict[str, Any]] = {}
        project_defs: list[dict[str, Any]] = []

        for row in parsed_rows:
            location_key: str | None = None
            if row.location_data is not None:
                normalized_location = self._normalize_location_payload(row.location_data)
                location_key = self._location_key(normalized_location)
                location_confidence = self._confidence_from_raw(row.raw, "location_confidence", 80)
                if location_key not in location_defs:
                    location_defs[location_key] = {
                        "normalized_data": normalized_location,
                        "raw": _sanitize_payload(row.raw),
                        "confidence": location_confidence,
                    }
                else:
                    existing_confidence = int(location_defs[location_key].get("confidence", 80))
                    location_defs[location_key]["confidence"] = max(
                        existing_confidence,
                        location_confidence,
                    )

            if row.project_data is not None:
                project_defs.append(
                    {
                        "normalized_data": self._normalize_project_payload(row.project_data),
                        "raw": _sanitize_payload(row.raw),
                        "location_key": location_key,
                        "confidence": self._confidence_from_raw(row.raw, "stream_confidence", 75),
                    }
                )

        items: list[ImportItem] = []
        location_items_by_key: dict[str, ImportItem] = {}
        candidate_location_ids_by_key: dict[str, list[UUID]] = {}

        for key, payload in location_defs.items():
            normalized = payload["normalized_data"]
            duplicate_candidates = self._find_location_duplicates(existing_locations, normalized)
            candidate_location_ids_by_key[key] = self._candidate_location_ids_for_key(
                location_key=key,
                duplicate_candidates=duplicate_candidates,
                existing_location_ids_by_key=existing_location_ids_by_key,
            )
            confidence = int(payload.get("confidence", 80))
            needs_review = (
                self._needs_review("location", normalized)
                or bool(duplicate_candidates)
                or self._needs_review_by_confidence(confidence)
            )
            item = ImportItem(
                organization_id=run.organization_id,
                run_id=run.id,
                item_type="location",
                status="pending_review",
                needs_review=needs_review,
                confidence=confidence,
                extracted_data=payload["raw"],
                normalized_data=normalized,
                duplicate_candidates=duplicate_candidates or None,
                confirm_create_new=False,
                group_id=self._group_id_for_location_key(run=run, location_key=key),
            )
            items.append(item)
            location_items_by_key[key] = item

        if items:
            db.add_all(items)
            await db.flush()

        all_candidate_location_ids: list[UUID] = []
        seen_location_ids: set[UUID] = set()
        for location_ids in candidate_location_ids_by_key.values():
            for location_id in location_ids:
                if location_id in seen_location_ids:
                    continue
                seen_location_ids.add(location_id)
                all_candidate_location_ids.append(location_id)
        project_index_by_location_and_name = await self._prefetch_project_index_for_locations(
            db=db,
            organization_id=run.organization_id,
            location_ids=all_candidate_location_ids,
        )

        # Build name-only lookup for location_ref fallback matching
        _location_by_name: dict[str, ImportItem] = {}
        for loc_item in location_items_by_key.values():
            loc_name = _normalize_token(
                str(loc_item.normalized_data.get("name") or "")
                if isinstance(loc_item.normalized_data, dict)
                else ""
            )
            if loc_name and loc_name not in _location_by_name:
                _location_by_name[loc_name] = loc_item

        for project_index, payload in enumerate(project_defs):
            normalized = payload["normalized_data"]
            location_key = payload["location_key"]
            parent_item = location_items_by_key.get(location_key) if location_key else None

            # Fallback: try matching stream_location_ref by name
            if parent_item is None:
                raw_ref = str(payload.get("raw", {}).get("stream_location_ref") or "")
                if raw_ref:
                    ref_name = _normalize_token(raw_ref)
                    parent_item = _location_by_name.get(ref_name)
                    if parent_item is not None:
                        # Update location_key so duplicate matching uses correct location_ids
                        location_key = self._location_key(
                            parent_item.normalized_data
                            if isinstance(parent_item.normalized_data, dict)
                            else {}
                        )

            if parent_item is None:
                confidence = int(payload.get("confidence", 50))
                item = ImportItem(
                    organization_id=run.organization_id,
                    run_id=run.id,
                    item_type="project",
                    status="pending_review",
                    needs_review=True,
                    confidence=confidence,
                    extracted_data=payload["raw"],
                    normalized_data=normalized,
                    review_notes=MISSING_LOCATION_REVIEW_NOTE,
                    confirm_create_new=False,
                    group_id=self._group_id_for_unresolved_project(
                        run=run,
                        project_index=project_index,
                    ),
                )
                items.append(item)
                continue

            location_ids = (
                candidate_location_ids_by_key.get(location_key, []) if location_key else []
            )
            duplicate_candidates = self._match_project_duplicates_from_index(
                project_index_by_location_and_name=project_index_by_location_and_name,
                location_ids=location_ids,
                project_data=normalized,
            )
            confidence = int(payload.get("confidence", 75))
            needs_review = (
                self._needs_review("project", normalized)
                or bool(duplicate_candidates)
                or self._needs_review_by_confidence(confidence)
            )
            item = ImportItem(
                organization_id=run.organization_id,
                run_id=run.id,
                item_type="project",
                status="pending_review",
                needs_review=needs_review,
                confidence=confidence,
                extracted_data=payload["raw"],
                normalized_data=normalized,
                duplicate_candidates=duplicate_candidates or None,
                parent_item_id=parent_item.id if parent_item else None,
                confirm_create_new=False,
                group_id=parent_item.group_id if parent_item else None,
            )
            items.append(item)

        return items

    async def _build_organization_entrypoint_items(
        self,
        db: AsyncSession,
        run: ImportRun,
        parsed_rows: list[ParsedRow],
    ) -> list[ImportItem]:
        location_defs: dict[str, dict[str, Any]] = {}
        project_defs: list[dict[str, Any]] = []

        for row in parsed_rows:
            location_key: str | None = None
            if row.location_data is not None:
                normalized_location = self._normalize_location_payload(row.location_data)
                location_key = self._location_key(normalized_location)
                location_confidence = self._confidence_from_raw(row.raw, "location_confidence", 80)
                if location_key not in location_defs:
                    location_defs[location_key] = {
                        "normalized_data": normalized_location,
                        "raw": _sanitize_payload(row.raw),
                        "confidence": location_confidence,
                    }
                else:
                    existing_confidence = int(location_defs[location_key].get("confidence", 80))
                    location_defs[location_key]["confidence"] = max(
                        existing_confidence,
                        location_confidence,
                    )

            if row.project_data is not None:
                project_defs.append(
                    {
                        "normalized_data": self._normalize_project_payload(row.project_data),
                        "raw": _sanitize_payload(row.raw),
                        "location_key": location_key,
                        "confidence": self._confidence_from_raw(row.raw, "stream_confidence", 75),
                    }
                )

        items: list[ImportItem] = []
        location_items_by_key: dict[str, ImportItem] = {}

        for key, payload in location_defs.items():
            normalized = payload["normalized_data"]
            confidence = int(payload.get("confidence", 80))
            needs_review = self._needs_review(
                "location", normalized
            ) or self._needs_review_by_confidence(confidence)
            item = ImportItem(
                organization_id=run.organization_id,
                run_id=run.id,
                item_type="location",
                status="pending_review",
                needs_review=needs_review,
                confidence=confidence,
                extracted_data=payload["raw"],
                normalized_data=normalized,
                duplicate_candidates=None,
                confirm_create_new=False,
                group_id=self._group_id_for_location_key(run=run, location_key=key),
            )
            items.append(item)
            location_items_by_key[key] = item

        if items:
            db.add_all(items)
            await db.flush()

        # Build name-only lookup for location_ref fallback matching.
        location_by_name: dict[str, ImportItem] = {}
        for loc_item in location_items_by_key.values():
            loc_name = _normalize_token(
                str(loc_item.normalized_data.get("name") or "")
                if isinstance(loc_item.normalized_data, dict)
                else ""
            )
            if loc_name and loc_name not in location_by_name:
                location_by_name[loc_name] = loc_item

        for project_index, payload in enumerate(project_defs):
            normalized = payload["normalized_data"]
            location_key = payload["location_key"]
            parent_item = location_items_by_key.get(location_key) if location_key else None

            # Fallback: try matching stream_location_ref by name.
            if parent_item is None:
                raw_ref = str(payload.get("raw", {}).get("stream_location_ref") or "")
                if raw_ref:
                    ref_name = _normalize_token(raw_ref)
                    parent_item = location_by_name.get(ref_name)

            if parent_item is None:
                confidence = int(payload.get("confidence", 50))
                item = ImportItem(
                    organization_id=run.organization_id,
                    run_id=run.id,
                    item_type="project",
                    status="pending_review",
                    needs_review=True,
                    confidence=confidence,
                    extracted_data=payload["raw"],
                    normalized_data=normalized,
                    review_notes=MISSING_LOCATION_REVIEW_NOTE,
                    confirm_create_new=False,
                    group_id=self._group_id_for_unresolved_project(
                        run=run,
                        project_index=project_index,
                    ),
                )
                items.append(item)
                continue

            confidence = int(payload.get("confidence", 75))
            needs_review = self._needs_review(
                "project", normalized
            ) or self._needs_review_by_confidence(confidence)
            item = ImportItem(
                organization_id=run.organization_id,
                run_id=run.id,
                item_type="project",
                status="pending_review",
                needs_review=needs_review,
                confidence=confidence,
                extracted_data=payload["raw"],
                normalized_data=normalized,
                duplicate_candidates=None,
                parent_item_id=parent_item.id,
                confirm_create_new=False,
                group_id=parent_item.group_id,
            )
            items.append(item)

        return items

    async def _build_location_entrypoint_items(
        self,
        db: AsyncSession,
        run: ImportRun,
        parsed_rows: list[ParsedRow],
    ) -> list[ImportItem]:
        location = await db.get(Location, run.entrypoint_id)
        if not location or location.organization_id != run.organization_id:
            raise ValueError("entrypoint_location_not_found")

        project_index_by_location_and_name = await self._prefetch_project_index_for_locations(
            db=db,
            organization_id=run.organization_id,
            location_ids=[location.id],
        )

        items: list[ImportItem] = []
        invalid_locations_by_key: dict[str, ImportItem] = {}

        for row in parsed_rows:
            if row.location_data:
                normalized_location = self._normalize_location_payload(row.location_data)
                key = self._location_key(normalized_location)
                if key and key not in invalid_locations_by_key:
                    confidence = self._confidence_from_raw(row.raw, "location_confidence", 50)
                    invalid_item = ImportItem(
                        organization_id=run.organization_id,
                        run_id=run.id,
                        item_type="location",
                        status="invalid",
                        needs_review=False,
                        confidence=confidence,
                        extracted_data=_sanitize_payload(row.raw),
                        normalized_data=normalized_location,
                        review_notes="Location items invalid for location entrypoint",
                        group_id=self._group_id_for_location_key(run=run, location_key=key),
                    )
                    invalid_locations_by_key[key] = invalid_item
                    items.append(invalid_item)

            if not row.project_data:
                continue

            normalized_project = self._normalize_project_payload(row.project_data)
            is_external_location = False
            if row.location_data:
                normalized_location = self._normalize_location_payload(row.location_data)
                is_external_location = self._is_external_location(location, normalized_location)

            duplicate_candidates = self._match_project_duplicates_from_index(
                project_index_by_location_and_name=project_index_by_location_and_name,
                location_ids=[location.id],
                project_data=normalized_project,
            )
            status_value = "invalid" if is_external_location else "pending_review"
            review_notes = (
                "Project row references external location" if is_external_location else None
            )
            confidence = self._confidence_from_raw(row.raw, "stream_confidence", 75)
            needs_review = (
                False
                if status_value == "invalid"
                else (
                    self._needs_review("project", normalized_project)
                    or bool(duplicate_candidates)
                    or self._needs_review_by_confidence(confidence)
                )
            )

            item = ImportItem(
                organization_id=run.organization_id,
                run_id=run.id,
                item_type="project",
                status=status_value,
                needs_review=needs_review,
                confidence=confidence,
                extracted_data=_sanitize_payload(row.raw),
                normalized_data=normalized_project,
                duplicate_candidates=duplicate_candidates or None,
                review_notes=review_notes,
                confirm_create_new=False,
                group_id=(
                    self._group_id_for_location_key(
                        run=run,
                        location_key=self._location_key(
                            self._normalize_location_payload(row.location_data)
                        ),
                    )
                    if row.location_data
                    else self._group_id_for_location_entrypoint(run)
                ),
            )
            items.append(item)

        if invalid_locations_by_key:
            db.add_all(list(invalid_locations_by_key.values()))
            await db.flush()

        return items

    def _normalize_location_payload(self, payload: dict[str, str]) -> dict[str, object]:
        normalized = {
            "name": _sanitize_text(payload.get("name") or "") or "",
            "city": _sanitize_text(payload.get("city") or "") or "",
            "state": _sanitize_text(payload.get("state") or "") or "",
            "address": _sanitize_text(payload.get("address") or ""),
        }
        return _sanitize_payload(normalized)

    def _normalize_project_payload(self, payload: dict[str, str]) -> dict[str, object]:
        normalized = {
            "name": _sanitize_text(payload.get("name") or "") or "",
            "category": _sanitize_text(payload.get("category") or ""),
            "project_type": _sanitize_text(payload.get("project_type") or "Assessment")
            or "Assessment",
            "description": _sanitize_text(payload.get("description") or ""),
            "sector": _sanitize_text(payload.get("sector") or ""),
            "subsector": _sanitize_text(payload.get("subsector") or ""),
            "estimated_volume": _sanitize_text(payload.get("estimated_volume") or ""),
            "volume": _sanitize_text(payload.get("volume") or ""),
            "units": _sanitize_text(payload.get("units") or ""),
            "frequency": _sanitize_text(payload.get("frequency") or ""),
            "company_name": _sanitize_text(payload.get("company_name") or "", max_len=255),
            "location_name": _sanitize_text(payload.get("location_name") or "", max_len=255),
            "location_city": _sanitize_text(payload.get("location_city") or "", max_len=100),
            "location_state": _sanitize_text(payload.get("location_state") or "", max_len=100),
            "location_address": _sanitize_text(payload.get("location_address") or "", max_len=500),
        }
        if not normalized["estimated_volume"] and (normalized["volume"] or normalized["frequency"]):
            volume_display = " ".join(
                part
                for part in [normalized["volume"], normalized["units"]]
                if isinstance(part, str) and part.strip()
            )
            estimated_parts = [
                part
                for part in [volume_display, normalized["frequency"]]
                if isinstance(part, str) and part.strip()
            ]
            normalized["estimated_volume"] = " / ".join(estimated_parts)
        return _sanitize_payload(normalized)

    def _location_key(self, normalized_location: dict[str, object]) -> str:
        name = _normalize_token(str(normalized_location.get("name") or ""))
        city = _normalize_token(str(normalized_location.get("city") or ""))
        state_value = _normalize_token(str(normalized_location.get("state") or ""))
        return f"{name}|{city}|{state_value}"

    def _group_id_for_location_key(self, *, run: ImportRun, location_key: str) -> str:
        digest = hashlib.sha256(f"{run.id}:{location_key}".encode()).hexdigest()
        return f"grp_{digest[:16]}"

    def _group_id_for_unresolved_project(
        self,
        *,
        run: ImportRun,
        project_index: int,
    ) -> str:
        digest = hashlib.sha256(f"{run.id}:unresolved:{project_index}".encode()).hexdigest()
        return f"grp_{digest[:16]}"

    def _group_id_for_location_entrypoint(self, run: ImportRun) -> str:
        digest = hashlib.sha256(f"{run.id}:entrypoint:{run.entrypoint_id}".encode()).hexdigest()
        return f"grp_{digest[:16]}"

    def _create_pending_intake_suggestions_from_project_item(
        self,
        *,
        db: AsyncSession,
        project: Project,
        source_item: ImportItem,
        registry: dict[str, Any],
        user_id: UUID,
    ) -> int:
        extracted = source_item.extracted_data
        raw_metadata = extracted.get("stream_metadata") if isinstance(extracted, dict) else None
        if not isinstance(raw_metadata, str) or not raw_metadata:
            return 0

        try:
            metadata = json.loads(raw_metadata)
        except json.JSONDecodeError:
            return 0
        if not isinstance(metadata, dict):
            return 0

        raw_hints = metadata.get("questionnaire_hints")
        if not isinstance(raw_hints, list):
            return 0

        created_count = 0
        seen_keys: set[tuple[str, str, str | None]] = set()
        for hint in raw_hints:
            if not isinstance(hint, dict):
                continue
            normalized_field_id = normalize_field_id(str(hint.get("field_id") or ""))
            registry_item_obj = registry.get(normalized_field_id)
            if registry_item_obj is None:
                continue

            if not hasattr(registry_item_obj, "section_id"):
                continue
            section_id = str(registry_item_obj.section_id)
            section_title = str(registry_item_obj.section_title)
            field_id = str(registry_item_obj.field_id)
            field_label = str(registry_item_obj.field_label)

            raw_value = hint.get("value")
            if raw_value is None:
                continue
            value = str(raw_value).strip()
            if not value:
                continue

            raw_unit = hint.get("unit")
            unit = str(raw_unit).strip() if isinstance(raw_unit, str) else None
            dedupe_key = (field_id, value, unit)
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)

            raw_confidence = hint.get("confidence")
            confidence_value = source_item.confidence if source_item.confidence is not None else 70
            if isinstance(raw_confidence, int):
                confidence_value = raw_confidence
            elif isinstance(raw_confidence, float):
                confidence_value = int(raw_confidence)
            confidence = max(0, min(100, confidence_value))

            value_type = "number" if self._is_number_like(value) else "string"

            suggestion = IntakeSuggestion(
                organization_id=project.organization_id,
                project_id=project.id,
                source_file_id=None,
                field_id=field_id,
                field_label=field_label,
                section_id=section_id,
                section_title=section_title,
                value=value,
                value_type=value_type,
                unit=unit,
                confidence=confidence,
                status="pending",
                source="notes",
                evidence=None,
                created_by_user_id=user_id,
            )
            created_count += 1
            db.add(suggestion)

        return created_count

    def _is_number_like(self, value: str) -> bool:
        try:
            float(value)
        except ValueError:
            return False
        return True

    def _find_location_duplicates(
        self,
        existing_locations: Sequence[Location],
        normalized_location: dict[str, object],
    ) -> list[dict[str, object]]:
        name = _normalize_token(str(normalized_location.get("name") or ""))
        city = _normalize_token(str(normalized_location.get("city") or ""))
        state_value = _normalize_token(str(normalized_location.get("state") or ""))
        if not name:
            return []

        candidates: list[dict[str, object]] = []
        for existing in existing_locations:
            existing_name = _normalize_token(existing.name)
            if existing_name != name:
                continue

            reasons = ["name_match"]
            if city and state_value:
                if (
                    _normalize_token(existing.city) == city
                    and _normalize_token(existing.state) == state_value
                ):
                    reasons.extend(["city_match", "state_match"])
                else:
                    continue
            candidates.append(
                {
                    "id": str(existing.id),
                    "name": existing.name,
                    "reason_codes": reasons,
                }
            )
        return candidates

    def _candidate_location_ids_for_key(
        self,
        *,
        location_key: str,
        duplicate_candidates: list[dict[str, object]],
        existing_location_ids_by_key: dict[str, list[UUID]],
    ) -> list[UUID]:
        candidate_ids: list[UUID] = []
        for candidate in duplicate_candidates:
            raw_id = candidate.get("id")
            if not isinstance(raw_id, str):
                continue
            try:
                candidate_ids.append(UUID(raw_id))
            except ValueError:
                continue

        if not candidate_ids:
            candidate_ids.extend(existing_location_ids_by_key.get(location_key, []))

        deduped: list[UUID] = []
        seen: set[UUID] = set()
        for candidate_id in candidate_ids:
            if candidate_id in seen:
                continue
            seen.add(candidate_id)
            deduped.append(candidate_id)
        return deduped

    async def _prefetch_project_index_for_locations(
        self,
        *,
        db: AsyncSession,
        organization_id: UUID,
        location_ids: Sequence[UUID],
    ) -> dict[tuple[UUID, str], list[Project]]:
        unique_location_ids = list(dict.fromkeys(location_ids))
        if not unique_location_ids:
            return {}

        result = await db.execute(
            select(Project).where(
                Project.organization_id == organization_id,
                Project.location_id.in_(unique_location_ids),
            )
        )
        projects = result.scalars().all()
        indexed: dict[tuple[UUID, str], list[Project]] = {}
        for project in projects:
            if project.location_id is None:
                continue
            normalized_name = _normalize_token(project.name)
            if not normalized_name:
                continue
            index_key = (project.location_id, normalized_name)
            if index_key not in indexed:
                indexed[index_key] = []
            indexed[index_key].append(project)
        return indexed

    def _match_project_duplicates_from_index(
        self,
        *,
        project_index_by_location_and_name: dict[tuple[UUID, str], list[Project]],
        location_ids: Sequence[UUID],
        project_data: dict[str, object],
    ) -> list[dict[str, object]]:
        project_name = _normalize_token(str(project_data.get("name") or ""))
        if not project_name:
            return []

        candidates: list[dict[str, object]] = []
        seen_project_ids: set[UUID] = set()
        for location_id in location_ids:
            matching_projects = project_index_by_location_and_name.get(
                (location_id, project_name), []
            )
            for project in matching_projects:
                if project.id in seen_project_ids:
                    continue
                seen_project_ids.add(project.id)
                candidates.append(
                    {
                        "id": str(project.id),
                        "name": project.name,
                        "reason_codes": ["name_match", "location_match"],
                    }
                )
        return candidates

    def _is_external_location(
        self, entrypoint_location: Location, candidate: dict[str, object]
    ) -> bool:
        candidate_name = _normalize_token(str(candidate.get("name") or ""))
        candidate_city = _normalize_token(str(candidate.get("city") or ""))
        candidate_state = _normalize_token(str(candidate.get("state") or ""))
        current_name = _normalize_token(entrypoint_location.name)
        current_city = _normalize_token(entrypoint_location.city)
        current_state = _normalize_token(entrypoint_location.state)

        if candidate_name and candidate_name != current_name:
            return True
        if candidate_city and candidate_city != current_city:
            return True
        return bool(candidate_state and candidate_state != current_state)
